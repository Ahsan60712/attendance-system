import os
import json
import pandas as pd
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

class WFHLeaveManager:
    def __init__(self, base_path):
        self.base_path = base_path
        excel_filename = os.environ.get('EXCEL_PATH', 'Emp_data.xlsx')
        self.emp_data_file = os.path.join(base_path, excel_filename)
    
    def authenticate_user(self, emp_name, password=None, role='employee'):
        """
        Authenticate user by emp_name and password
        role: 'employee' or 'admin'
        """
        try:
            if not os.path.exists(self.emp_data_file):
                raise FileNotFoundError(f"Employee data file not found: {self.emp_data_file}")
            
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            
            # Find employee by name
            emp_row = df[df['emp_name'].str.lower() == emp_name.lower()]
            
            if emp_row.empty:
                return None
            
            emp_data = emp_row.iloc[0].to_dict()
            
            # Verify password if provided (skip check if password column missing for backward compatibility)
            if password and 'password' in emp_data:
                # Convert both to string for comparison to handle Excel number formatting
                if str(emp_data['password']) != str(password):
                    return None
            
            # Check role requirements
            if role == 'admin' and not emp_data.get('is_admin', 0):
                return None
            if role == 'manager' and not emp_data.get('is_manager', 0) and not emp_data.get('is_admin', 0):
                # Admins can also act as managers
                return None
            
            # Return full employee data including counts
            return emp_data
            
        except Exception as e:
            print(f"Error authenticating user: {str(e)}")
            raise Exception(f"Authentication failed. Error: {str(e)}")

    def change_password(self, emp_id, new_password):
        """Update employee password"""
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            
            # Convert emp_id to proper type if needed
            df.loc[df['emp_id'] == int(emp_id), 'password'] = str(new_password)
            
            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
            return True
            
        except Exception as e:
            print(f"Error changing password: {str(e)}")
            raise Exception(f"Could not change password. Please ensure Emp_data.xlsx is closed. Error: {str(e)}")
    
    def get_employees(self):
        """Read employee list from Emp_data.xlsx"""
        try:
            if not os.path.exists(self.emp_data_file):
                raise FileNotFoundError(f"Employee data file not found: {self.emp_data_file}")
            
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            return df.to_dict('records')
        except Exception as e:
            print(f"Error reading employee data: {str(e)}")
            raise Exception(f"Could not read employee data. Please ensure Emp_data.xlsx is closed in Excel and is a valid Excel file. Error: {str(e)}")
    
    def get_daily_filepath(self, date_obj, create_dir=False):
        """Construct path: base_path/YYYY/Month/dd-mon-yyyy.xlsx"""
        year = date_obj.strftime('%Y')
        month = date_obj.strftime('%B') # Full month name
        filename = date_obj.strftime('%d-%b-%Y').lower() + '.xlsx'
        
        folder_path = os.path.join(self.base_path, year, month)
        
        if create_dir and not os.path.exists(folder_path):
            os.makedirs(folder_path)
            
        return os.path.join(folder_path, filename)

    # ===== CONTRACT YEAR LEAVE SYSTEM =====

    def get_contract_year_window(self, contract_start_date_str, today=None):
        """
        Given a contract start date string (YYYY-MM-DD), return the
        (year_start, year_end) of the CURRENT contract year that contains today.
        e.g. contract start = 2024-05-15, today = 2026-03-11
             → year 2 started 2025-05-15, ends 2026-05-14
             → returns (date(2025,5,15), date(2026,5,14))
        """
        if today is None:
            today = date.today()
        try:
            contract_start = date.fromisoformat(str(contract_start_date_str).split(' ')[0].split('T')[0])
        except Exception:
            return None, None

        # Walk forward by years until we pass today
        year_start = contract_start
        while True:
            year_end = year_start + relativedelta(years=1) - timedelta(days=1)
            if year_start <= today <= year_end:
                return year_start, year_end
            if year_start > today:
                # Shouldn't happen but guard against infinite loop
                break
            year_start = year_start + relativedelta(years=1)
        return None, None

    def check_and_rollover_leaves(self, emp_id):
        """
        Check if the employee's contract year has rolled over since the last
        time we ran. If yes:
          1. Carry forward unused leaves (up to max = Total_leaves cap or unlimited)
          2. Reset Leaves_This_Year to 0
          3. Recalculate Remaining_Leaves = Total_leaves + Leaves_Carried_Forward
          4. Store the new Contract_Year_Start
        Returns the updated employee row dict, or None on error.
        """
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')

            # Ensure new columns exist
            for col, default in [('Leaves_This_Year', 0), ('Leaves_Carried_Forward', 0), ('Contract_Year_Start', '')]:
                if col not in df.columns:
                    df[col] = default

            emp_row = df[df['emp_id'] == int(emp_id)]
            if emp_row.empty:
                return None
            idx = emp_row.index[0]

            contract_start_str = df.at[idx, 'Contract_Start_Date']
            if not contract_start_str or pd.isna(contract_start_str):
                # No contract date set — nothing to roll over
                return df.iloc[idx].to_dict()

            today = date.today()
            year_start, year_end = self.get_contract_year_window(contract_start_str, today)
            if year_start is None:
                return df.iloc[idx].to_dict()

            year_start_str = year_start.strftime('%Y-%m-%d')

            # Read stored year start
            stored_year_start = df.at[idx, 'Contract_Year_Start']
            if pd.isna(stored_year_start) or str(stored_year_start).strip() == '':
                stored_year_start = None

            if stored_year_start == year_start_str:
                # Already on the correct contract year — no rollover needed
                return df.iloc[idx].to_dict()

            # ---- ROLLOVER ----
            # Leaves taken in the OLD year = Leaves_This_Year
            old_leaves_taken = float(df.at[idx, 'Leaves_This_Year']) if pd.notna(df.at[idx, 'Leaves_This_Year']) else 0
            old_total = float(df.at[idx, 'Total_leaves']) if pd.notna(df.at[idx, 'Total_leaves']) else 0
            old_carried = float(df.at[idx, 'Leaves_Carried_Forward']) if pd.notna(df.at[idx, 'Leaves_Carried_Forward']) else 0

            # Unused in old year = (Total + carried) - taken  (can't go below 0)
            old_remaining = max(0, (old_total + old_carried) - old_leaves_taken)

            # Carry forward unused leaves (no cap — adjust if you want a cap e.g. min(old_remaining, 7))
            new_carried_forward = old_remaining

            # Reset for new year
            df.at[idx, 'Leaves_This_Year'] = 0
            df.at[idx, 'Leaves_Carried_Forward'] = new_carried_forward
            df.at[idx, 'Contract_Year_Start'] = year_start_str
            # Remaining = fresh allocation + carried
            df.at[idx, 'Remaining_Leaves'] = old_total + new_carried_forward
            # Reset cumulative Leaves counter to 0 for new year
            df.at[idx, 'Leaves'] = 0
            df.at[idx, 'Half_Day'] = 0

            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
            print(f"[ROLLOVER] Employee {emp_id}: carried forward {new_carried_forward} leaves into new year starting {year_start_str}")
            return df.iloc[idx].to_dict()

        except Exception as e:
            print(f"Error in check_and_rollover_leaves: {str(e)}")
            return None

    def get_leave_balance_info(self, emp_id):
        """
        Return a dict with rich leave balance details for display:
        - contract_year_start, contract_year_end (formatted strings)
        - total_leaves (annual entitlement)
        - carried_forward (from last year)
        - total_available (total + carried)
        - leaves_taken_this_year
        - half_days_taken
        - remaining_leaves
        """
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            for col, default in [('Leaves_This_Year', 0), ('Leaves_Carried_Forward', 0), ('Contract_Year_Start', '')]:
                if col not in df.columns:
                    df[col] = default

            emp_row = df[df['emp_id'] == int(emp_id)]
            if emp_row.empty:
                return {}
            row = emp_row.iloc[0]

            contract_start_str = row.get('Contract_Start_Date')
            year_start, year_end = None, None
            if contract_start_str and not pd.isna(contract_start_str):
                year_start, year_end = self.get_contract_year_window(str(contract_start_str))

            total = float(row.get('Total_leaves') or 0)
            carried = float(row.get('Leaves_Carried_Forward') or 0)
            taken_this_year = float(row.get('Leaves_This_Year') or 0)
            half_days = float(row.get('Half_Day') or 0)
            remaining = float(row.get('Remaining_Leaves') or 0)

            total_available = total + carried
            # Real-time calculation to avoid discrepancies with Excel stored value
            calculated_remaining = total_available - taken_this_year

            return {
                'contract_year_start': year_start.strftime('%d %b %Y') if year_start else 'N/A',
                'contract_year_end': year_end.strftime('%d %b %Y') if year_end else 'N/A',
                'total_leaves': total,
                'carried_forward': carried,
                'total_available': total_available,
                'leaves_taken_this_year': taken_this_year,
                'half_days_taken': half_days,
                'remaining_leaves': calculated_remaining,
            }
        except Exception as e:
            print(f"Error in get_leave_balance_info: {str(e)}")
            return {}
    
    def mark_wfh_leave(self, emp_id, emp_name, emp_team, date, request_type, reason, status='Pending', manager_name=''):
        """
        Mark WFH or Leave for an employee
        request_type: 'WFH' or 'Leave'
        reason: mandatory text explaining the request
        status: 'Pending' or 'Approved'
        manager_name: Name of manager if auto-approved
        """
        # Get structured filepath
        workbook_path = self.get_daily_filepath(date, create_dir=True)
        
        # Delete existing file if it exists (to ensure fresh data)
        # REMOVED: This was deleting the file and causing previous entries to be lost
        # We will instead load the workbook if it exists
        
        # Get all employees
        employees = self.get_employees()
        
        if os.path.exists(workbook_path):
            wb = load_workbook(workbook_path)
            ws = wb.active
            # Fix: If old file has no Status/Action By headers, add them now
            if ws.cell(row=1, column=7).value is None:
                ws.cell(row=1, column=7, value='Status')
                ws.cell(row=1, column=8, value='Action By')
                # Backfill existing data rows with Pending status
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value and ws.cell(row=r, column=7).value is None:
                        ws.cell(row=r, column=7, value='Pending')
                        ws.cell(row=r, column=8, value='')
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = date.strftime('%d-%b-%Y').lower()
            
            # Header styling
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            # Headers with Approval Status
            headers = ['Emp ID', 'Employee Name', 'Team', 'Request Type', 'Reason', 'Timestamp', 'Status', 'Action By']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Find the next available row
        next_row = ws.max_row + 1

        # Add only the employee's request
        ws.cell(row=next_row, column=1, value=emp_id)
        ws.cell(row=next_row, column=2, value=emp_name)
        ws.cell(row=next_row, column=3, value=emp_team)
        
        # Request type with color coding
        type_cell = ws.cell(row=next_row, column=4, value=request_type)
        if request_type == 'WFH':
            type_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            type_cell.font = Font(color='9C6500')
        elif request_type == 'Leave':
            type_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            type_cell.font = Font(color='9C0006')
        elif request_type == 'Half Day':
            type_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            type_cell.font = Font(color='38761D')
            
        type_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Reason, Timestamp, Status
        ws.cell(row=next_row, column=5, value=reason)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=next_row, column=6, value=timestamp)
        ws.cell(row=next_row, column=7, value=status)
        ws.cell(row=next_row, column=8, value=manager_name if status == 'Approved' else '')
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        
        # Save workbook
        wb.save(workbook_path)
        
        # If auto-approved (manager case), update counters and log for admin immediately
        if status == 'Approved':
            self.update_employee_counters(emp_id, request_type)
            self.log_approval_action(
                emp_id=emp_id,
                emp_name=emp_name,
                request_type=request_type,
                request_date=date.strftime('%Y-%m-%d'),
                status='Approved',
                manager_name=manager_name or emp_name
            )
        
        return True
    
    def update_employee_counters(self, emp_id, request_type):
        """Update WFH_count, Leaves, or Half_Day counter in Emp_data.xlsx"""
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')

            # Ensure contract-year columns exist
            for col, default in [('Leaves_This_Year', 0), ('Leaves_Carried_Forward', 0), ('Contract_Year_Start', '')]:
                if col not in df.columns:
                    df[col] = default
            
            emp_row = df[df['emp_id'] == int(emp_id)]
            if emp_row.empty:
                return
            
            idx = emp_row.index[0]
            
            if request_type == 'WFH':
                df.at[idx, 'WFH_count'] = df.at[idx, 'WFH_count'] + 1 if pd.notna(df.at[idx, 'WFH_count']) else 1
            elif request_type == 'Leave':
                if pd.notna(df.at[idx, 'Remaining_Leaves']) and df.at[idx, 'Remaining_Leaves'] < 1:
                    raise Exception("Insufficient remaining leaves")
                df.at[idx, 'Leaves'] = df.at[idx, 'Leaves'] + 1 if pd.notna(df.at[idx, 'Leaves']) else 1
                # Also track this-year counter
                df.at[idx, 'Leaves_This_Year'] = (df.at[idx, 'Leaves_This_Year'] + 1) if pd.notna(df.at[idx, 'Leaves_This_Year']) else 1
                if pd.notna(df.at[idx, 'Remaining_Leaves']):
                    df.at[idx, 'Remaining_Leaves'] = df.at[idx, 'Remaining_Leaves'] - 1
            elif request_type == 'Half Day':
                if pd.notna(df.at[idx, 'Remaining_Leaves']) and df.at[idx, 'Remaining_Leaves'] < 0.5:
                    raise Exception("Insufficient remaining leaves for a half day")
                df.at[idx, 'Half_Day'] = df.at[idx, 'Half_Day'] + 1 if pd.notna(df.at[idx, 'Half_Day']) else 1
                # Half day = 0.5 leave this year
                df.at[idx, 'Leaves_This_Year'] = (df.at[idx, 'Leaves_This_Year'] + 0.5) if pd.notna(df.at[idx, 'Leaves_This_Year']) else 0.5
                if pd.notna(df.at[idx, 'Remaining_Leaves']):
                    df.at[idx, 'Remaining_Leaves'] = df.at[idx, 'Remaining_Leaves'] - 0.5
            
            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
            
        except Exception as e:
            print(f"Error updating employee counters: {str(e)}")
            raise Exception(f"Could not update Emp_data.xlsx. Please ensure it is closed in Excel. Error: {str(e)}")

    def refund_employee_counters(self, emp_id, request_type):
        """Reverse the counters if an approved request is cancelled."""
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            emp_row = df[df['emp_id'] == int(emp_id)]
            if emp_row.empty:
                return
            idx = emp_row.index[0]
            
            if request_type == 'WFH':
                df.at[idx, 'WFH_count'] = max((df.at[idx, 'WFH_count'] - 1) if pd.notna(df.at[idx, 'WFH_count']) else 0, 0)
            elif request_type == 'Leave':
                df.at[idx, 'Leaves'] = max((df.at[idx, 'Leaves'] - 1) if pd.notna(df.at[idx, 'Leaves']) else 0, 0)
                df.at[idx, 'Leaves_This_Year'] = max((df.at[idx, 'Leaves_This_Year'] - 1) if pd.notna(df.at[idx, 'Leaves_This_Year']) else 0, 0)
                if pd.notna(df.at[idx, 'Remaining_Leaves']):
                    df.at[idx, 'Remaining_Leaves'] = df.at[idx, 'Remaining_Leaves'] + 1
            elif request_type == 'Half Day':
                df.at[idx, 'Half_Day'] = max((df.at[idx, 'Half_Day'] - 1) if pd.notna(df.at[idx, 'Half_Day']) else 0, 0)
                df.at[idx, 'Leaves_This_Year'] = max((df.at[idx, 'Leaves_This_Year'] - 0.5) if pd.notna(df.at[idx, 'Leaves_This_Year']) else 0, 0)
                if pd.notna(df.at[idx, 'Remaining_Leaves']):
                    df.at[idx, 'Remaining_Leaves'] = df.at[idx, 'Remaining_Leaves'] + 0.5
            
            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
        except Exception as e:
            print(f"Error refunding counters: {str(e)}")
            raise Exception(f"Could not refund Emp_data.xlsx. Error: {str(e)}")
    

    def get_notifications(self, filter_date=None, limit=None):
        """
        Get WFH/Leave requests for a specific date (default: today)
        If limit is provided, it's ignored in this new date-based logic 
        (we show all requests for the day)
        """
        notifications = []
        
        try:
            if filter_date is None:
                filter_date = date.today()
            
            # Use helper to get structured path
            filepath = self.get_daily_filepath(filter_date)
            
            if os.path.exists(filepath):
                try:
                    wb = load_workbook(filepath, data_only=True)
                    ws = wb.active
                    
                    # Read each row (skip header)
                    for row_num in range(2, ws.max_row + 1):
                        emp_id = ws.cell(row=row_num, column=1).value
                        emp_name = ws.cell(row=row_num, column=2).value
                        team = ws.cell(row=row_num, column=3).value
                        request_type = ws.cell(row=row_num, column=4).value
                        reason = ws.cell(row=row_num, column=5).value
                        timestamp = ws.cell(row=row_num, column=6).value
                        status = ws.cell(row=row_num, column=7).value
                        
                        # Set default status for older files without the status column
                        if status is None:
                            status = 'Approved' 
                            
                        if emp_id:
                            notifications.append({
                                'date': filter_date.strftime('%d-%b-%Y'),
                                'emp_id': emp_id,
                                'emp_name': emp_name,
                                'team': team,
                                'type': request_type,
                                'reason': reason,
                                'timestamp': timestamp,
                                'status': status
                            })
                            
                    # Reverse to show newest first for that day
                    notifications.reverse()
                    
                except Exception as e:
                    print(f"Error reading notification file {filepath}: {str(e)}")
            
            return notifications
            
        except Exception as e:
            print(f"Error getting notifications: {str(e)}")
            return []

    def get_pending_requests(self, days_back=30, days_forward=365, req_status='Pending'):
        """Scan recent and future daily files for requests based on req_status"""
        requests_list = []
        try:
            from datetime import timedelta
            end_date = date.today() + timedelta(days=days_forward)
            start_date = date.today() - timedelta(days=days_back)
            
            current_date = start_date
            while current_date <= end_date:
                filepath = self.get_daily_filepath(current_date)
                
                if os.path.exists(filepath):
                    try:
                        wb = load_workbook(filepath, data_only=True)
                        ws = wb.active
                        
                        for row_num in range(2, ws.max_row + 1):
                            emp_id = ws.cell(row=row_num, column=1).value
                            if not emp_id:
                                continue
                            status = ws.cell(row=row_num, column=7).value
                            if req_status == 'Pending':
                                match = status in ('Pending', None)
                            else:
                                match = status == req_status
                                
                            if match:
                                emp_name = ws.cell(row=row_num, column=2).value
                                team = ws.cell(row=row_num, column=3).value
                                request_type = ws.cell(row=row_num, column=4).value
                                reason = ws.cell(row=row_num, column=5).value
                                timestamp = ws.cell(row=row_num, column=6).value
                                
                                requests_list.append({
                                    'date': current_date.strftime('%Y-%m-%d'),
                                    'display_date': current_date.strftime('%d-%b-%Y'),
                                    'emp_id': emp_id,
                                    'emp_name': emp_name,
                                    'team': team,
                                    'type': request_type,
                                    'reason': reason,
                                    'timestamp': timestamp,
                                    'status': status
                                })
                    except Exception as e:
                        print(f"Error reading {filepath}: {str(e)}")
                
                current_date += timedelta(days=1)
                
            # Sort newest first
            requests_list.sort(key=lambda x: x['timestamp'], reverse=True)
            return requests_list
            
        except Exception as e:
            print(f"Error getting pending requests: {str(e)}")
            return []
            
    def update_request_status(self, request_date_str, emp_id, request_type, timestamp, new_status, manager_name):
        """Update a specific request from Pending to Approved or Rejected"""
        try:
            req_date = datetime.strptime(request_date_str, '%Y-%m-%d').date()
            filepath = self.get_daily_filepath(req_date)
            
            if not os.path.exists(filepath):
                raise Exception(f"File not found for date: {request_date_str}")
                
            wb = load_workbook(filepath)
            ws = wb.active
            
            updated = False
            emp_name_found = ''
            for row_num in range(2, ws.max_row + 1):
                file_emp_id = ws.cell(row=row_num, column=1).value
                file_timestamp = ws.cell(row=row_num, column=6).value
                
                # Match exactly the right row using employee ID and their unique timestamp
                if str(file_emp_id) == str(emp_id) and str(file_timestamp) == str(timestamp):
                    emp_name_found = ws.cell(row=row_num, column=2).value or ''
                    ws.cell(row=row_num, column=7, value=new_status)
                    ws.cell(row=row_num, column=8, value=manager_name)
                    updated = True
                    break
                    
            if not updated:
                raise Exception("Could not find the specific request to update.")
                
            wb.save(filepath)
            
            # If approved, deduct from employee counters
            if new_status == 'Approved':
                self.update_employee_counters(emp_id, request_type)
            
            # Log the action for admin notifications
            self.log_approval_action(
                emp_id=emp_id,
                emp_name=emp_name_found,
                request_type=request_type,
                request_date=request_date_str,
                status=new_status,
                manager_name=manager_name
            )
                
            return True
        except Exception as e:
            print(f"Error updating request status: {str(e)}")
            raise Exception(f"Failed to update status: {str(e)}")

    def cancel_request(self, request_date_str, emp_id, timestamp, cancelled_by):
        """Cancel an existing request (pending or approved), refund counters if it was approved."""
        try:
            req_date = datetime.strptime(request_date_str, '%Y-%m-%d').date()
            filepath = self.get_daily_filepath(req_date)
            
            if not os.path.exists(filepath):
                raise Exception(f"File not found for date: {request_date_str}")
                
            wb = load_workbook(filepath)
            ws = wb.active
            
            updated = False
            for row_num in range(2, ws.max_row + 1):
                file_emp_id = ws.cell(row=row_num, column=1).value
                file_timestamp = ws.cell(row=row_num, column=6).value
                
                if str(file_emp_id) == str(emp_id) and str(file_timestamp) == str(timestamp):
                    current_status = ws.cell(row=row_num, column=7).value
                    request_type = ws.cell(row=row_num, column=4).value
                    
                    if current_status == 'Approved':
                        self.refund_employee_counters(emp_id, request_type)
                        
                    ws.cell(row=row_num, column=7, value='Cancelled')
                    ws.cell(row=row_num, column=8, value=cancelled_by)
                    updated = True
                    break
                    
            if not updated:
                raise Exception("Could not find the specific request to cancel.")
                
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Error cancelling request: {str(e)}")
            raise Exception(f"Failed to cancel request: {str(e)}")

    def log_approval_action(self, emp_id, emp_name, request_type, request_date, status, manager_name):
        """Append an approved/rejected action entry to approved_log.json"""
        log_path = os.path.join(self.base_path, 'approved_log.json')
        try:
            if os.path.exists(log_path):
                with open(log_path, 'r', encoding='utf-8') as f:
                    log = json.load(f)
            else:
                log = []
            
            log.insert(0, {
                'emp_id': emp_id,
                'emp_name': emp_name,
                'request_type': request_type,
                'request_date': request_date,
                'status': status,
                'manager_name': manager_name,
                'actioned_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Keep only the latest 200 entries
            log = log[:200]
            
            with open(log_path, 'w', encoding='utf-8') as f:
                json.dump(log, f, indent=2)
        except Exception as e:
            print(f"Error writing approval log: {str(e)}")

    def get_approval_log(self, limit=50):
        """Return recent approved/rejected actions for admin notifications"""
        log_path = os.path.join(self.base_path, 'approved_log.json')
        try:
            if not os.path.exists(log_path):
                return []
            with open(log_path, 'r', encoding='utf-8') as f:
                log = json.load(f)
            return log[:limit]
        except Exception as e:
            print(f"Error reading approval log: {str(e)}")
            return []

    def get_employee_records(self, emp_id, start_date, end_date):
        """Get all WFH/Leave records for an employee within date range"""
        records = []
        
        try:
            current_date = start_date
            while current_date <= end_date:
                # Use helper to get structured path
                filepath = self.get_daily_filepath(current_date)
                
                if os.path.exists(filepath):
                    try:
                        wb = load_workbook(filepath, data_only=True)
                        ws = wb.active
                        
                        for row_num in range(2, ws.max_row + 1):
                            file_emp_id = ws.cell(row=row_num, column=1).value
                            
                            # Determine if this row belongs to the employee
                            if str(file_emp_id) == str(emp_id):
                                status = ws.cell(row=row_num, column=7).value
                                if status is None:
                                    status = 'Approved' # Legacy default
                                    
                                records.append({
                                    'date': current_date.strftime('%d-%b-%Y'),
                                    'raw_date': current_date.strftime('%Y-%m-%d'),
                                    'type': ws.cell(row=row_num, column=4).value,
                                    'reason': ws.cell(row=row_num, column=5).value,
                                    'timestamp': ws.cell(row=row_num, column=6).value,
                                    'status': status
                                })
                    except:
                        pass
                
                # Move to next day
                from datetime import timedelta
                current_date += timedelta(days=1)
            
            return records
            
        except Exception as e:
            print(f"Error getting employee records: {str(e)}")
            return []

    def add_employee(self, emp_name, emp_team, is_admin, is_manager, contract_type, contract_start_date, contract_end_date, total_leaves, password='SecurePass2026!'):
        """Add a new employee to Emp_data.xlsx"""
        try:
            if not os.path.exists(self.emp_data_file):
                raise FileNotFoundError("Employee data file not found")
            
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')

            # Ensure contract-year columns exist
            for col, default in [('Leaves_This_Year', 0), ('Leaves_Carried_Forward', 0), ('Contract_Year_Start', '')]:
                if col not in df.columns:
                    df[col] = default
            
            # Generate new emp_id (max + 1)
            new_id = int(df['emp_id'].max() + 1) if not df.empty else 1

            # Compute the initial contract-year start for display
            year_start_str = ''
            if contract_start_date:
                try:
                    ys, _ = self.get_contract_year_window(contract_start_date)
                    if ys:
                        year_start_str = ys.strftime('%Y-%m-%d')
                except Exception:
                    pass
            
            new_emp = {
                'emp_id': new_id,
                'emp_name': emp_name,
                'emp_team': emp_team,
                'Total_leaves': total_leaves,
                'Half_Day': 0,
                'Leaves': 0,
                'Leaves_This_Year': 0,
                'Leaves_Carried_Forward': 0,
                'Present': 0,
                'Remaining_Leaves': total_leaves,
                'is_admin': 1 if is_admin else 0,
                'is_manager': 1 if is_manager else 0,
                'WFH_count': 0,
                'password': password,
                'Contract_Type': contract_type,
                'Contract_Start_Date': contract_start_date,
                'Contract_End_Date': contract_end_date,
                'Contract_Year_Start': year_start_str,
            }
            
            # Append new row
            df = pd.concat([df, pd.DataFrame([new_emp])], ignore_index=True)
            
            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
            return new_id
            
        except Exception as e:
            print(f"Error adding employee: {str(e)}")
            raise Exception(f"Could not add employee: {str(e)}")

    def update_employee(self, emp_id, emp_name, emp_team, is_admin, is_manager, contract_type, contract_start_date, contract_end_date, total_leaves, carried_forward=None, phone=None):
        """Update an existing employee in Emp_data.xlsx"""
        try:
            if not os.path.exists(self.emp_data_file):
                raise FileNotFoundError("Employee data file not found")
            
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            
            # Ensure contract-year columns exist
            for col, default in [('Leaves_This_Year', 0), ('Leaves_Carried_Forward', 0), ('Contract_Year_Start', '')]:
                if col not in df.columns:
                    df[col] = default

            # Find the row index
            emp_idx = df.index[df['emp_id'] == int(emp_id)].tolist()
            if not emp_idx:
                raise Exception(f"Employee with ID {emp_id} not found")
            
            idx = emp_idx[0]
            
            # Prevent dtype crashing if empty excel columns were read as float64
            for col in ['Contract_Type', 'Contract_Start_Date', 'Contract_End_Date', 'Contract_Year_Start']:
                if col in df.columns and df[col].dtype == 'float64':
                    df[col] = df[col].astype(object)
            
            df.at[idx, 'emp_name'] = str(emp_name) if emp_name else df.at[idx, 'emp_name']
            df.at[idx, 'emp_team'] = str(emp_team) if emp_team else df.at[idx, 'emp_team']
            df.at[idx, 'is_admin'] = 1 if is_admin else 0
            df.at[idx, 'is_manager'] = 1 if is_manager else 0
            df.at[idx, 'Contract_Type'] = str(contract_type) if contract_type else df.at[idx, 'Contract_Type']
            
            # If contract start date changed, we should reset the internal 'Contract_Year_Start'
            # trigger so the system re-calculates the current contract year and rollover.
            old_csd = str(df.at[idx, 'Contract_Start_Date']).strip()
            new_csd = str(contract_start_date).strip() if contract_start_date else ''
            
            if contract_start_date: 
                df.at[idx, 'Contract_Start_Date'] = new_csd
                if new_csd != old_csd:
                    # Date changed -> clear the calc'd year start to trigger re-rollover check
                    df.at[idx, 'Contract_Year_Start'] = ''

            if contract_end_date: df.at[idx, 'Contract_End_Date'] = str(contract_end_date)
            
            # Manual update of carried forward leaves
            if carried_forward is not None:
                df.at[idx, 'Leaves_Carried_Forward'] = float(carried_forward)

            # Recalculate remaining leaves
            # Remaining = (New Total + New Carried) - (Leaves Taken This Year)
            old_total = df.at[idx, 'Total_leaves']
            df.at[idx, 'Total_leaves'] = float(total_leaves)
            
            taken_this_year = float(df.at[idx, 'Leaves_This_Year']) if pd.notna(df.at[idx, 'Leaves_This_Year']) else 0
            current_carried = float(df.at[idx, 'Leaves_Carried_Forward']) if pd.notna(df.at[idx, 'Leaves_Carried_Forward']) else 0
            
            # Ensure phone column exists
            if 'phone' not in df.columns:
                df['phone'] = ''

            if 'phone' in df.columns and df['phone'].dtype == 'float64':
                df['phone'] = df['phone'].astype(object)

            df.at[idx, 'Remaining_Leaves'] = (float(total_leaves) + current_carried) - taken_this_year

            # Save phone number if provided
            if phone is not None:
                df.at[idx, 'phone'] = str(phone).strip()

            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
            return True
            
        except Exception as e:
            print(f"Error updating employee: {str(e)}")
            raise Exception(f"Could not update employee: {str(e)}")

    def delete_employee(self, emp_id):
        """Delete an employee from Emp_data.xlsx"""
        try:
            if not os.path.exists(self.emp_data_file):
                raise FileNotFoundError("Employee data file not found")
            
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            
            # Remove employee row
            df = df[df['emp_id'] != int(emp_id)]
            
            df.to_excel(self.emp_data_file, index=False, engine='openpyxl')
            return True
            
        except Exception as e:
            print(f"Error deleting employee: {str(e)}")
            raise Exception(f"Could not delete employee: {str(e)}")

    # ── WhatsApp / notification helpers ──────────────────────────────────────

    def get_manager_for_team(self, team_name: str) -> dict:
        """
        Return the manager row for the given team as a dict, or {}.
        The dict includes 'emp_name', 'phone' (if set), etc.
        If the team has no dedicated manager the first admin is returned as fallback.
        """
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            # Ensure phone column exists
            if 'phone' not in df.columns:
                df['phone'] = ''

            # 1. Look for a manager in the same team
            team_managers = df[
                (df['emp_team'].str.strip().str.lower() == str(team_name).strip().lower()) &
                (df['is_manager'] == 1)
            ]
            if not team_managers.empty:
                return team_managers.iloc[0].to_dict()

            # 2. Fallback: any admin
            admins = df[df['is_admin'] == 1]
            if not admins.empty:
                return admins.iloc[0].to_dict()

            return {}
        except Exception as e:
            print(f"Error in get_manager_for_team: {e}")
            return {}

    def get_employee_phone(self, emp_id) -> str:
        """Return the phone number stored for an employee, or empty string."""
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            if 'phone' not in df.columns:
                return ''
            emp_row = df[df['emp_id'] == int(emp_id)]
            if emp_row.empty:
                return ''
            phone = emp_row.iloc[0].get('phone', '')
            return str(phone) if phone and not pd.isna(phone) else ''
        except Exception as e:
            print(f"Error in get_employee_phone: {e}")
            return ''

    def get_all_managers(self) -> list:
        """Return list of dicts for all managers (is_manager=1) with phone numbers."""
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            if 'phone' not in df.columns:
                df['phone'] = ''
            managers = df[df['is_manager'] == 1]
            return managers.to_dict('records')
        except Exception as e:
            print(f"Error in get_all_managers: {e}")
            return []

    def get_daily_attendance_summary(self, summary_date):
        """
        Build a per-team attendance summary for the given date.
        Returns: {
            'team_name': {
                'wfh': [emp_name, ...],
                'leave': [emp_name, ...],
                'half_day': [emp_name, ...],
                'no_request': [emp_name, ...]   # employees with no entry for the day
            }, ...
        }
        """
        try:
            df = pd.read_excel(self.emp_data_file, engine='openpyxl')
            filepath = self.get_daily_filepath(summary_date)

            # Build a set of emp_ids who filed something for this date
            filed_requests = {}   # emp_id → {'type': ..., 'status': ...}
            if os.path.exists(filepath):
                wb = load_workbook(filepath, data_only=True)
                ws = wb.active
                for row_num in range(2, ws.max_row + 1):
                    emp_id = ws.cell(row=row_num, column=1).value
                    if not emp_id:
                        continue
                    req_type = ws.cell(row=row_num, column=4).value
                    status = ws.cell(row=row_num, column=7).value or 'Approved'
                    # Only count non-cancelled / non-rejected requests
                    if status in ('Approved', 'Pending'):
                        filed_requests[str(emp_id)] = {
                            'type': req_type,
                            'status': status
                        }

            # Group employees by team
            teams = {}
            for _, row in df.iterrows():
                team = str(row.get('emp_team', 'Unknown')).strip()
                emp_id = str(row.get('emp_id', ''))
                emp_name = str(row.get('emp_name', ''))
                is_manager = row.get('is_manager', 0)
                is_admin = row.get('is_admin', 0)

                if team not in teams:
                    teams[team] = {'wfh': [], 'leave': [], 'half_day': [], 'no_request': []}

                if emp_id in filed_requests:
                    req = filed_requests[emp_id]
                    rtype = req['type']
                    if rtype == 'WFH':
                        teams[team]['wfh'].append(emp_name)
                    elif rtype == 'Leave':
                        teams[team]['leave'].append(emp_name)
                    elif rtype == 'Half Day':
                        teams[team]['half_day'].append(emp_name)
                # If no request filed and it's a weekday, mark as no_request
                # (Skip managers/admins from the "no request" list — they self-approve)
                elif summary_date.weekday() < 5:
                    # Don't flag managers/admins as absent
                    if not is_manager and not is_admin:
                        teams[team]['no_request'].append(emp_name)

            return teams
        except Exception as e:
            print(f"Error in get_daily_attendance_summary: {e}")
            return {}

