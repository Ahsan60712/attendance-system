
import snowflake.connector
import pandas as pd
import os
from datetime import datetime, date
from openpyxl import load_workbook

# --- CONFIGURATION ---
SNOW_ACCOUNT  = 'wxlwerb-rg07665'
SNOW_USER     = 'MUHAMMAD.AHSAN'     
SNOW_PASSWORD = 'Ahsan123$' 
SNOW_DATABASE = 'ADLABS'
SNOW_SCHEMA   = 'AHSAN'
SNOW_WH       = 'ADLABS_WH'
SNOW_ROLE     = 'ADLABS_ROLE'

BASE_PATH = r'c:\Users\ahsan\Desktop\Attandance management system'
EMP_FILE = os.path.join(BASE_PATH, 'Emp_data.xlsx')

def get_connection():
    return snowflake.connector.connect(
        user=SNOW_USER,
        password=SNOW_PASSWORD,
        account=SNOW_ACCOUNT,
        warehouse=SNOW_WH,
        database=SNOW_DATABASE,
        schema=SNOW_SCHEMA,
        role=SNOW_ROLE
    )

def migrate_employees():
    print("--- 1. Migrating Employees ---")
    df = pd.read_excel(EMP_FILE)
    conn = get_connection()
    cursor = conn.cursor()
    
    for _, row in df.iterrows():
        try:
            sql = f"""INSERT INTO ADLABS.AHSAN.EMPLOYEES (
                EMP_ID, EMP_NAME, EMP_TEAM, PASSWORD, IS_ADMIN, IS_MANAGER, 
                CONTRACT_START_DATE, TOTAL_LEAVES, REMAINING_LEAVES, 
                LEAVES_THIS_YEAR, LEAVES_CARRIED_FORWARD, PHONE
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            cursor.execute(sql, (
                int(row['emp_id']), str(row['emp_name']), str(row.get('emp_team', 'General')),
                str(row.get('password', 'SecurePass2026!')), bool(row.get('is_admin', 0)),
                bool(row.get('is_manager', 0)), str(row.get('contract_start_date', '2024-01-01')),
                float(row.get('Total_leaves', 14)), float(row.get('Remaining_Leaves', 14)),
                float(row.get('Leaves_This_Year', 0)), float(row.get('Leaves_Carried_Forward', 0)),
                str(row.get('phone', ''))
            ))
        except Exception as e:
            print(f"  Skipped {row.get('emp_name')}: {e}")
            
    conn.commit()
    cursor.close()
    conn.close()
    print("Employees Uploaded!")

def migrate_attendance_history():
    print("--- 2. Migrating Attendance History ---")
    conn = get_connection()
    cursor = conn.cursor()
    
    # Traverse directory (YYYY/Month format)
    for root, dirs, files in os.walk(BASE_PATH):
        for file in files:
            if file.endswith('.xlsx') and file != 'Emp_data.xlsx' and '-' in file:
                try:
                    # Parse date from filename (e.g. 15-may-2024.xlsx)
                    date_str = file.replace('.xlsx', '')
                    req_date = datetime.strptime(date_str, '%d-%b-%Y').date()
                    
                    filepath = os.path.join(root, file)
                    wb = load_workbook(filepath, data_only=True)
                    ws = wb.active
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0]: continue
                        
                        sql = "INSERT INTO ADLABS.AHSAN.ATTENDANCE_REQUESTS (EMP_ID, REQUEST_DATE, REQUEST_TYPE, REASON, STATUS, APPROVED_BY, SUBMITTED_AT) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                        cursor.execute(sql, (
                            row[0], req_date.strftime('%Y-%m-%d'), row[3], row[4], 
                            row[6] if len(row)>6 else 'Approved', 
                            row[7] if len(row)>7 else 'System',
                            row[5] if len(row)>5 else datetime.now()
                        ))
                    print(f"  Processed: {file}")
                except Exception as e:
                    print(f"  Error in {file}: {e}")
                    
    conn.commit()
    cursor.close()
    conn.close()
    print("Attendance Migration Complete!")

if __name__ == "__main__":
    migrate_employees()
    migrate_attendance_history()
