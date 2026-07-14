from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from attendance_manager import WFHLeaveManager
from datetime import date, datetime, timedelta
import os
import logging
from dotenv import load_dotenv
import threading

# Load environment variables from .env
load_dotenv()

import whatsapp_notifier as wa

logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'

# Initialize the WFH/Leave manager (Ab yeh pure database se link hona chahiye)
BASE_PATH = os.path.dirname(os.path.abspath(__file__))
manager = WFHLeaveManager(BASE_PATH)

def _is_empty_or_nan(val):
    """Helper function to replace pandas pd.isna check for database values"""
    if val is None:
        return True
    s_val = str(val).strip().lower()
    return s_val in ['', 'none', 'nan', 'nat', '<na>']

def _attach_request_leave_balance(req, balance_cache):
    """Set emp_balance from ATTENDANCE_REQUESTS-based calculation from Database."""
    emp_id = req.get('emp_id')
    if not emp_id:
        req['emp_balance'] = 'N/A'
        return
    bal = manager.get_leave_balance_cached(emp_id, balance_cache)
    req['emp_balance'] = bal.get('remaining_leaves', 'N/A')
    req['total_allotted'] = bal.get('total_available', bal.get('total_leaves', 14))

def _enrich_employees_leave_balances(employees):
    """Balances are already populated during get_employees() mapping. No-op to save 20+ database queries."""
    pass

def _sync_all_balances_on_startup():
    try:
        manager.sync_all_employee_counters()
    except Exception as e:
        print(f"[Startup] Leave balance sync failed: {e}")

# threading.Thread(target=_sync_all_balances_on_startup, daemon=True).start()

# ── APScheduler — Automated Daily Summary & Absent Alert ─────────────────────
from apscheduler.schedulers.background import BackgroundScheduler
import atexit

@app.route('/privacy-policy')
def privacy_policy():
    """Privacy policy page for Meta API compliance"""
    return render_template('privacy_policy.html')

def _run_daily_leave_checks():
    """Background job: Automatically check and apply rollover/expiry from DB daily."""
    print("[Scheduler] ▶ Running daily leave expiry and rollover checks for all employees...")
    try:
        employees = manager.get_employees()
        for emp in employees:
            emp_id = emp.get('emp_id')
            if emp_id:
                manager.check_and_rollover_leaves(emp_id)
                manager.check_and_apply_expiry(emp_id)
        # manager.sync_all_employee_counters()  # Disabled to prevent overwriting manually entered data in EMPLOYEES table
        print("[Scheduler] ✅ Daily leave checks completed for all employees.")
    except Exception as e:
        print(f"[Scheduler] ❌ Error in daily leave checks: {e}")

def _send_daily_summaries():
    """Background job: build per-team attendance summary from DB and WhatsApp each manager."""
    today = date.today()
    # Skip weekends
    if today.weekday() >= 5:
        print(f"[Scheduler] Skipping daily summary — it's a weekend ({today}).")
        return

    print(f"[Scheduler] ▶ Sending daily attendance summaries for {today}…")
    try:
        teams = manager.get_daily_attendance_summary(today)
        all_managers = manager.get_all_managers()

        for mgr in all_managers:
            mgr_phone = mgr.get('phone', '')
            mgr_name = mgr.get('emp_name', 'Manager')
            mgr_team = str(mgr.get('emp_team', '')).strip()

            if _is_empty_or_nan(mgr_phone):
                print(f"[Scheduler] Skipping {mgr_name} — no phone number.")
                continue

            team_data = teams.get(mgr_team, {})
            wa.send_daily_summary(
                manager_phone=str(mgr_phone),
                manager_name=mgr_name,
                summary_date=today,
                team_name=mgr_team or 'All',
                wfh_list=team_data.get('wfh', []),
                leave_list=team_data.get('leave', []),
                half_day_list=team_data.get('half_day', []),
                absent_list=team_data.get('no_request', [])
            )
        print(f"[Scheduler] ✅ Daily summaries sent for {today}.")
    except Exception as e:
        print(f"[Scheduler] ❌ Error sending daily summaries: {e}")


def _send_absent_alerts():
    """Background job: alert each manager about employees with no request filed."""
    today = date.today()
    if today.weekday() >= 5:
        print(f"[Scheduler] Skipping absent alert — it's a weekend ({today}).")
        return

    print(f"[Scheduler] ▶ Sending absent-employee alerts for {today}…")
    try:
        teams = manager.get_daily_attendance_summary(today)
        all_managers = manager.get_all_managers()

        for mgr in all_managers:
            mgr_phone = mgr.get('phone', '')
            mgr_name = mgr.get('emp_name', 'Manager')
            mgr_team = str(mgr.get('emp_team', '')).strip()

            if _is_empty_or_nan(mgr_phone):
                continue

            team_data = teams.get(mgr_team, {})
            absent = team_data.get('no_request', [])
            if absent:
                # ── Notify Manager ──
                wa.send_absent_alert(
                    manager_phone=str(mgr_phone),
                    manager_name=mgr_name,
                    absent_employees=absent,
                    team_name=mgr_team or 'All',
                    alert_date=today
                )

                # ── Notify Each Absent Employee Directly ──
                employees = manager.get_employees()
                for emp_name in absent:
                    emp_info = next((e for e in employees if str(e.get('emp_name')).strip() == str(emp_name).strip()), {})
                    emp_phone = emp_info.get('phone', '')
                    if not _is_empty_or_nan(emp_phone):
                        wa.notify_employee_absence(
                            emp_phone=str(emp_phone),
                            emp_name=emp_name,
                            alert_date=today
                        )
        print(f"[Scheduler] ✅ Absent alerts sent for {today}.")
    except Exception as e:
        print(f"[Scheduler] ❌ Error sending absent alerts: {e}")


def _run_annual_contract_renewal():
    """Scheduler job: runs on July 1st each year — advances all employee contract dates by 1 year."""
    today = date.today()
    print(f"[Scheduler] ▶ Running annual contract renewal for {today}...")
    try:
        results = manager.renew_all_contracts()
        renewed  = [r for r in results if r['status'] == 'renewed']
        skipped  = [r for r in results if r['status'] == 'skipped']
        errors   = [r for r in results if r['status'] == 'error']
        for r in renewed:
            print(f"[Scheduler]   ✅ {r['emp']}: {r['old_start']} → {r['old_end']}  ⟹  {r['new_start']} → {r['new_end']}")
        for r in skipped:
            print(f"[Scheduler]   ⏭  {r['emp']}: skipped — {r['reason']}")
        for r in errors:
            print(f"[Scheduler]   ❌ {r['emp']}: ERROR — {r['reason']}")
        print(f"[Scheduler] ✅ Contract renewal done — {len(renewed)} renewed, {len(skipped)} skipped, {len(errors)} errors.")
    except Exception as e:
        print(f"[Scheduler] ❌ Error in annual contract renewal: {e}")

SUMMARY_HOUR = int(os.environ.get('DAILY_SUMMARY_HOUR', 9))
SUMMARY_MINUTE = int(os.environ.get('DAILY_SUMMARY_MINUTE', 30))
ALERT_HOUR = int(os.environ.get('ABSENT_ALERT_HOUR', 11))
ALERT_MINUTE = int(os.environ.get('DAILY_SUMMARY_MINUTE', 0))

scheduler = BackgroundScheduler(daemon=True)

scheduler.add_job(
    _run_daily_leave_checks,
    trigger='cron',
    hour=0,
    minute=5,
    id='daily_leave_checks',
    name='Daily Leave Expiry and Rollover'
)

scheduler.add_job(
    _send_daily_summaries,
    trigger='cron',
    hour=SUMMARY_HOUR,
    minute=SUMMARY_MINUTE,
    day_of_week='mon-fri',
    id='daily_summary',
    name='Daily Attendance Summary'
)
scheduler.add_job(
    _send_absent_alerts,
    trigger='cron',
    hour=ALERT_HOUR,
    minute=ALERT_MINUTE,
    day_of_week='mon-fri',
    id='absent_alert',
    name='Absent Employee Alert'
)
scheduler.add_job(
    _run_annual_contract_renewal,
    trigger='cron',
    month=7,
    day=1,
    hour=0,
    minute=1,
    id='annual_contract_renewal',
    name='Annual Contract Renewal (Jul 1)'
)

scheduler.start()
atexit.register(lambda: scheduler.shutdown(wait=False))
print(f"[Scheduler] ✅ Started — Leave checks at 00:05, Daily summary at {SUMMARY_HOUR:02d}:{SUMMARY_MINUTE:02d}, Absent alert at {ALERT_HOUR:02d}:{ALERT_MINUTE:02d}")

@app.route('/')
def index():
    return render_template('index.html')

# ===== EMPLOYEE ROUTES =====

@app.route('/employee-login', methods=['GET', 'POST'])
def employee_login():
    if request.method == 'POST':
        emp_name = request.form.get('emp_name')
        password = request.form.get('password')
        
        try:
            user = manager.authenticate_user(emp_name, password)
            if user:
                if user.get('is_manager', 0) == 1:
                    flash('Access denied. Team Leads must log in through the Manager Portal.', 'error')
                    return redirect(url_for('employee_login'))

                session.pop('_flashes', None)
                session['logged_in'] = True
                flash('Login successful', 'success')

                session['user_type'] = 'employee'
                session['emp_id'] = user['emp_id']
                session['emp_name'] = user['emp_name']
                session['emp_team'] = user.get('emp_team', '')
                return redirect(url_for('employee_dashboard'))
            else:
                flash('Wrong credentials', 'error')
        except Exception as e:
            flash(f'Login failed: {str(e)}', 'error')
    
    return render_template('employee_login.html')

def _dashboard_url_for_user_type(user_type):
    if user_type == 'admin':
        return url_for('admin_dashboard')
    if user_type == 'ceo':
        return url_for('ceo_dashboard')
    if user_type == 'manager':
        return url_for('manager_dashboard')
    return url_for('employee_dashboard')

@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    if not session.get('logged_in') or not session.get('emp_id'):
        return redirect(url_for('index'))

    user_type = session.get('user_type')
    if user_type not in ('employee', 'manager', 'admin', 'ceo'):
        return redirect(url_for('index'))

    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not current_password or not new_password or not confirm_password:
            flash('All fields are required.', 'error')
        elif new_password != confirm_password:
            flash('New password and confirm password do not match.', 'error')
        elif len(new_password) < 4:
            flash('New password must be at least 4 characters.', 'error')
        elif new_password == current_password:
            flash('New password must be different from your current password.', 'error')
        else:
            try:
                manager.change_password(session['emp_id'], current_password, new_password)
                flash('Your password has been updated successfully.', 'success')
                return redirect(_dashboard_url_for_user_type(user_type))
            except Exception as e:
                flash(str(e), 'error')

    return render_template('change_password.html')

@app.route('/employee-dashboard')
def employee_dashboard():
    if session.get('user_type') != 'employee':
        return redirect(url_for('employee_login'))
    
    emp_id = session.get('emp_id')
    today = date.today()
    
    manager.check_and_rollover_leaves(emp_id)
    manager.check_and_apply_expiry(emp_id)

    employees = manager.get_employees()
    current_emp = next((e for e in employees if str(e['emp_id']) == str(emp_id)), {})

    leave_balance = manager.get_leave_balance_info(emp_id)
    filter_type = request.args.get('filter_type', 'default')
    
    if filter_type == 'monthly':
        month = int(request.args.get('month', today.month))
        year = int(request.args.get('year', today.year))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
            
    elif filter_type == 'yearly':
        year = int(request.args.get('year', today.year))
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
    elif filter_type == 'custom':
        start_str = request.args.get('start_date')
        end_str = request.args.get('end_date')
        if start_str and end_str:
            start_date = date.fromisoformat(start_str)
            end_date = date.fromisoformat(end_str)
        else:
            start_date = today - timedelta(days=30)
            end_date = today + timedelta(days=365)
            
    else:
        start_date = today - timedelta(days=30)
        end_date = today + timedelta(days=365)

    records = manager.get_employee_records(emp_id, start_date, end_date)
    
    return render_template('employee_dashboard.html', 
                         emp_name=session.get('emp_name'),
                         emp_data=current_emp,
                         leave_balance=leave_balance,
                         records=records,
                         today=today.strftime('%Y-%m-%d'),
                         filter_type=filter_type,
                         filter_start=start_date.strftime('%Y-%m-%d'),
                         filter_end=end_date.strftime('%Y-%m-%d'))

@app.route('/manager-mark-attendance', methods=['POST'])
def manager_mark_attendance():
    if session.get('user_type') != 'manager':
        return redirect(url_for('manager_login'))
    
    try:
        target_emp_id = request.form.get('target_emp_id')
        request_type = request.form.get('request_type')
        reason = request.form.get('reason')
        start_date_str = request.form.get('date', date.today().strftime('%Y-%m-%d'))
        end_date_str = request.form.get('end_date')
        
        if not target_emp_id:
            flash('Please select an employee', 'error')
            return redirect(url_for('manager_dashboard'))
        
        if not reason or not reason.strip():
            flash('Reason is required', 'error')
            return redirect(url_for('manager_dashboard'))
        
        target_emp = manager.get_employee_by_id(target_emp_id)
        if not target_emp:
            flash('Employee not found', 'error')
            return redirect(url_for('manager_dashboard'))
        
        start_date = date.fromisoformat(start_date_str)
        
        if end_date_str and len(end_date_str.strip()) > 5:
            try:
                end_date = date.fromisoformat(end_date_str)
                if end_date < start_date:
                    end_date = start_date
            except:
                end_date = start_date
        else:
            end_date = start_date
        
        dates_to_mark = []
        current = start_date
        while current <= end_date:
            dates_to_mark.append(current)
            current += timedelta(days=1)
        
        manager_name = session.get('emp_name')
        
        for d in dates_to_mark:
            manager.mark_wfh_leave(
                emp_id=target_emp_id,
                emp_name=target_emp.get('emp_name'),
                emp_team=target_emp.get('emp_team'),
                date=d,
                request_type=request_type,
                reason=reason,
                status='Approved',
                manager_name=manager_name
            )
        
        flash(f'Successfully marked {request_type} for {target_emp.get("emp_name")} from {start_date.strftime("%d %b %Y")} to {end_date.strftime("%d %b %Y")}', 'success')
        return redirect(url_for('manager_dashboard'))
        
    except Exception as e:
        flash(f'Error marking attendance: {str(e)}', 'error')
        return redirect(url_for('manager_dashboard'))

@app.route('/mark-request', methods=['POST'])
def mark_request():
    user_type = session.get('user_type')
    if user_type not in ['employee', 'manager']:
        return redirect(url_for('employee_login'))
    
    status = 'Approved' if user_type == 'manager' else 'Pending'
    manager_name = session.get('emp_name') if user_type == 'manager' else ''
    redirect_target = 'manager_dashboard' if user_type == 'manager' else 'employee_dashboard'
    
    try:
        request_type = request.form.get('request_type')
        reason = request.form.get('reason')
        start_date_str = request.form.get('date', date.today().strftime('%Y-%m-%d'))
        end_date_str = request.form.get('end_date')
        
        if not reason or not reason.strip():
            flash('Reason is required', 'error')
            return redirect(url_for(redirect_target))
        
        start_date = date.fromisoformat(start_date_str)
        
        from datetime import timezone
        today = (datetime.now(timezone.utc) + timedelta(hours=5)).date()
        
        max_days_back = 2 if user_type == 'employee' else 0
        min_allowed_date = today - timedelta(days=max_days_back)
        
        if start_date < min_allowed_date:
            if user_type == 'employee':
                flash(f'Error: You cannot apply for dates more than 2 days back. Today is {today.strftime("%d %b %Y")}.', 'error')
            else:
                flash(f'Error: Invalid date selected.', 'error')
            return redirect(url_for(redirect_target))
            
        if end_date_str and len(end_date_str.strip()) > 5:
            try:
                end_date = date.fromisoformat(end_date_str)
                if end_date < start_date:
                    end_date = start_date
            except:
                end_date = start_date
        else:
            end_date = start_date

        dates_to_mark = []
        current = start_date
        while current <= end_date:
            dates_to_mark.append(current)
            current += timedelta(days=1)
        
        for d in dates_to_mark:
            manager.mark_wfh_leave(
                emp_id=session.get('emp_id'),
                emp_name=session.get('emp_name'),
                emp_team=session.get('emp_team'),
                date=d,
                request_type=request_type,
                reason=reason,
                status=status,
                manager_name=manager_name
            )
            success_msg = f'{request_type} applied successfully!'

        if user_type == 'employee':
            try:
                emp_team = session.get('emp_team', '')
                team_manager = manager.get_manager_for_team(emp_team)
                mgr_phone = team_manager.get('phone', '')
                wa.notify_manager_new_request(
                    manager_phone=mgr_phone or '',
                    manager_name=team_manager.get('emp_name', 'Manager'),
                    emp_name=session.get('emp_name', ''),
                    request_type=request_type,
                    request_date=f"{start_date.strftime('%d %b %Y')} – {end_date.strftime('%d %b %Y')}" if end_date != start_date else start_date.strftime('%d %b %Y'),
                    reason=reason
                )
                
                emp_id = session.get('emp_id')
                emp_phone = manager.get_employee_phone(emp_id)
                if emp_phone:
                    wa.notify_employee_request_submitted(
                        emp_phone=emp_phone,
                        emp_name=session.get('emp_name', ''),
                        request_type=request_type,
                        request_date=f"{start_date.strftime('%d %b %Y')} – {end_date.strftime('%d %b %Y')}" if end_date != start_date else start_date.strftime('%d %b %Y')
                    )
            except Exception as wa_err:
                flash(f"WhatsApp Error: {str(wa_err)}", 'danger')

        flash(success_msg, 'success')
        return redirect(url_for(redirect_target))
        
    except Exception as e:
        flash(f'Error submitting request: {str(e)}', 'error')
        return redirect(url_for(redirect_target))

# ===== MANAGER ROUTES =====

@app.route('/manager-login', methods=['GET', 'POST'])
def manager_login():
    if request.method == 'POST':
        emp_name = request.form.get('emp_name')
        password = request.form.get('password')
        
        try:
            user = manager.authenticate_user(emp_name, password, role='manager')
            if user:
                session['logged_in'] = True
                session['user_type'] = 'manager'
                session['emp_id'] = user['emp_id']
                session['emp_name'] = user['emp_name']
                session['emp_team'] = user['emp_team']
                flash('Login successful', 'success')
                return redirect(url_for('manager_dashboard'))
            else:
                flash('Wrong credentials', 'error')
        except Exception as e:
            flash(f'Login failed: {str(e)}', 'error')
    
    return render_template('manager_login.html')

@app.route('/manager-dashboard')
def manager_dashboard():
    try:
        if session.get('user_type') not in ['manager', 'admin']:
            return redirect(url_for('employee_login'))
        
        emp_id = session.get('emp_id')
        emp_name = session.get('emp_name', '')
        today = date.today()

        start_date_str = request.args.get('start_date', (today - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date_str = request.args.get('end_date', today.strftime('%Y-%m-%d'))

        manager.check_and_rollover_leaves(emp_id)
        manager.check_and_apply_expiry(emp_id)
        
        employees = manager.get_employees()
        current_emp = next((e for e in employees if str(e.get('emp_id')) == str(emp_id)), {})
        
        if not current_emp:
            current_emp = {
                'emp_id': emp_id,
                'emp_name': session.get('emp_name', 'Manager'),
                'emp_team': session.get('emp_team', 'General')
            }

        leave_balance = manager.get_leave_balance_info(emp_id)
        is_sajeel = 'sajeel' in emp_name.lower()
        
        all_pending_requests = manager.get_pending_requests(req_status='Pending')
        manager_team = (current_emp.get('emp_team') or session.get('emp_team', '')).strip().lower()
        
        pending_requests = []
        leave_balance_cache = {}
        for req in all_pending_requests:
            req_emp_id = req.get('emp_id')
            if str(req_emp_id) == str(emp_id):
                continue
            
            req_emp_data = next((e for e in employees if str(e.get('emp_id')) == str(req_emp_id)), {})
            req_emp_team = (req_emp_data.get('emp_team') or '').strip().lower()
            
            if is_sajeel:
                is_target_team = req_emp_team in ['poppi', 'ovg']
            else:
                is_target_team = (req_emp_team == manager_team)
            
            if is_target_team:
                _attach_request_leave_balance(req, leave_balance_cache)
                pending_requests.append(req)
                
        all_approved_requests = manager.get_pending_requests(req_status='Approved')
        team_approved = []
        for req in all_approved_requests:
            req_emp_id = req.get('emp_id')
            if str(req_emp_id) == str(emp_id):
                continue
            
            req_emp_data = next((e for e in employees if str(e.get('emp_id')) == str(req_emp_id)), {})
            req_emp_team = (req_emp_data.get('emp_team') or '').strip().lower()
            
            if is_sajeel:
                is_target_team = req_emp_team in ['poppi', 'ovg']
            else:
                is_target_team = (req_emp_team == manager_team)
            
            if is_target_team:
                team_approved.append(req)

        approved_requests = []
        try:
            sd = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            ed = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            for req in team_approved:
                req_date_raw = req.get('date')
                if req_date_raw:
                    if isinstance(req_date_raw, (date, datetime)):
                        req_date = req_date_raw if isinstance(req_date_raw, date) else req_date_raw.date()
                    else:
                        req_date = datetime.strptime(str(req_date_raw).split(' ')[0], '%Y-%m-%d').date()
                    if sd <= req_date <= ed:
                        approved_requests.append(req)
        except Exception as e:
            approved_requests = team_approved

        my_records = manager.get_employee_records(emp_id, start_date_str, end_date_str)
        
        if not isinstance(leave_balance, dict):
            leave_balance = {}
        
        overstock_members = manager.get_overstock_team_members()
        
        if is_sajeel:
            team_members = [e for e in employees if (e.get('emp_team') or '').strip().lower() in ['poppi', 'ovg']]
            team_members_balances = [e for e in employees if (e.get('emp_team') or '').strip().lower() in ['poppi', 'ovg']]
        else:
            team_members = [e for e in employees if (e.get('emp_team') or '').strip().lower() == manager_team]
            team_members_balances = [e for e in employees if (e.get('emp_team') or '').strip().lower() == manager_team]
            
        team_members_balances.sort(key=lambda e: e.get('emp_name', '').lower())
        
        schedules = manager.get_shift_schedules()
        if schedules:
            latest_submitted_schedule = max(schedules, key=lambda s: s.get('submitted_at') or s.get('valid_from'))
            latest_date = latest_submitted_schedule.get('valid_from')
            if latest_date:
                schedules = [s for s in schedules if s.get('valid_from') == latest_date]
        
        emp_dict = {str(e.get('emp_id')): e.get('emp_name') for e in manager.get_employees()}
        # Invert schedules to find who is assigned to what main shift
        emp_to_shift = {}
        meeting_lead_week1 = 'Not Assigned'
        meeting_lead_week2 = 'Not Assigned'
        weekly_report_week1 = 'Not Assigned'
        weekly_report_week2 = 'Not Assigned'
        
        for s in schedules:
            emp_id_str = str(s.get('emp_id'))
            schedule_type = s.get('schedule_type')
            shift_name = s.get('shift_name')
            
            # Resolve names for weekly roles
            resolved_name = next((e.get('emp_name') for e in employees if str(e.get('emp_id')) == emp_id_str), 'Not Assigned')
            
            if schedule_type == 'meeting_lead_week1':
                meeting_lead_week1 = resolved_name
            elif schedule_type == 'meeting_lead_week2':
                meeting_lead_week2 = resolved_name
            elif schedule_type == 'weekly_report_week1':
                weekly_report_week1 = resolved_name
            elif schedule_type == 'weekly_report_week2':
                weekly_report_week2 = resolved_name
            elif schedule_type == 'main':
                emp_to_shift[emp_id_str] = shift_name
                
        # Shift schedules definitions
        SHIFT_SCHEDULES = {
            'Weekend Night': {
                'MON': 'Off Day', 'TUE': 'Off Day', 'WED': 'Weekend Night', 'THU': 'Weekend Night', 
                'FRI': 'Off Day', 'SAT': 'Weekend Night', 'SUN': 'Weekend Night'
            },
            'Weekend Morning': {
                'MON': 'Weekend Morning', 'TUE': 'Weekend Morning', 'WED': 'Off Day', 'THU': 'Off Day', 
                'FRI': 'Off Day', 'SAT': 'Weekend Morning', 'SUN': 'Weekend Morning'
            },
            'Night': {
                'MON': 'Night', 'TUE': 'Night', 'WED': 'Night', 'THU': 'Night', 
                'FRI': 'Night', 'SAT': 'Off Day', 'SUN': 'Off Day'
            },
            'Morning': {
                'MON': 'Morning', 'TUE': 'Morning', 'WED': 'Morning', 'THU': 'Morning', 
                'FRI': 'Morning', 'SAT': 'Off Day', 'SUN': 'Off Day'
            },
            'Primary (P1)': {
                'MON': 'Primary (P1)', 'TUE': 'Primary (P1)', 'WED': 'Primary (P1)', 'THU': 'Primary (P1)', 
                'FRI': 'Primary (P1)', 'SAT': 'Off Day', 'SUN': 'Off Day'
            },
            'Primary (P2)': {
                'MON': 'Primary (P2)', 'TUE': 'Primary (P2)', 'WED': 'Primary (P2)', 'THU': 'Primary (P2)', 
                'FRI': 'Primary (P2)', 'SAT': 'Off Day', 'SUN': 'Off Day'
            },
            'Development Office': {
                'MON': 'Dev 1', 'TUE': 'Dev 1', 'WED': 'Dev 1', 'THU': 'Dev 1', 
                'FRI': 'Dev 1', 'SAT': 'Off Day', 'SUN': 'Off Day'
            },
            'Development office': {
                'MON': 'Dev 2', 'TUE': 'Dev 2', 'WED': 'Dev 2', 'THU': 'Dev 2', 
                'FRI': 'Dev 2', 'SAT': 'Off Day', 'SUN': 'Off Day'
            }
        }
        
        SHIFT_DISPLAY = {
            'Weekend Night': 'Weekend Night',
            'Weekend Morning': 'Weekend Morning',
            'Night': 'Night',
            'Morning': 'Morning',
            'Primary (P1)': 'Primary (P1)',
            'Primary (P2)': 'Primary (P2)',
            'Development Office': 'Dev 1',
            'Development office': 'Dev 2'
        }
        
        # Filter: only display Overstock team members (employees and managers)
        overstock_employees = [e for e in employees if (e.get('emp_team') or '').strip().lower() == 'overstock']
        # Sort managers to the top, then by name
        overstock_employees.sort(key=lambda e: (not e.get('is_manager', False), e.get('emp_name', '').lower()))
        
        grid_data = []
        for emp in overstock_employees:
            emp_id_str = str(emp.get('emp_id'))
            emp_name = emp.get('emp_name')
            is_manager = emp.get('is_manager', False)
            
            assigned_shift = emp_to_shift.get(emp_id_str)
            
            days_schedule = {}
            for d in ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']:
                if assigned_shift and assigned_shift in SHIFT_SCHEDULES:
                    days_schedule[d] = SHIFT_SCHEDULES[assigned_shift][d]
                else:
                    days_schedule[d] = 'Off Day'
                    
            grid_data.append({
                'emp_id': emp.get('emp_id'),
                'emp_name': emp_name,
                'is_manager': is_manager,
                'assigned_shift': assigned_shift or 'Not Assigned',
                'assigned_shift_display': SHIFT_DISPLAY.get(assigned_shift, 'Not Assigned') if assigned_shift else 'Not Assigned',
                'days': days_schedule
            })
        
        return render_template('manager_dashboard.html', 
                               manager_name=session.get('emp_name', 'Manager'),
                               emp_data=current_emp or {},
                               leave_balance=leave_balance,
                               requests=pending_requests,
                               approved_requests=approved_requests,
                               my_records=my_records,
                               start_date=start_date_str,
                               end_date=end_date_str,
                               total_allotted=leave_balance.get('total_allotted', 0),
                               carried_forward=leave_balance.get('carried_forward', 0),
                               remaining_leaves=leave_balance.get('remaining_leaves', 0),
                               today=today.strftime('%Y-%m-%d'),
                               overstock_members=overstock_members,
                               grid_data=grid_data,
                               meeting_lead_week1=meeting_lead_week1,
                               meeting_lead_week2=meeting_lead_week2,
                               weekly_report_week1=weekly_report_week1,
                               weekly_report_week2=weekly_report_week2,
                               team_members=team_members,
                               team_members_balances=team_members_balances)
    except Exception as e:
        return f"<h2>⚠️ Dashboard Crash!</h2><p>Error Type: {type(e).__name__}</p><p>Message: {str(e)}</p>"

@app.template_filter('format_number')
def format_number_filter(value):
    try:
        f_val = float(value)
        return int(f_val) if f_val.is_integer() else f_val
    except:
        return value

@app.route('/action-request', methods=['POST'])
def action_request():
    if session.get('user_type') not in ['manager', 'admin']:
        if request.headers.get('Content-Type') == 'application/json' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        return redirect(url_for('employee_login'))
        
    emp_id = request.form.get('emp_id')
    req_date = request.form.get('date')
    req_type = request.form.get('type')
    timestamp = request.form.get('timestamp')
    action = request.form.get('action')
    
    try:
        new_status = 'Approved' if action == 'Approve' else 'Rejected'
        manager.update_request_status(
            request_date_str=req_date, 
            emp_id=emp_id, 
            request_type=req_type,
            timestamp=timestamp, 
            new_status=new_status, 
            manager_name=session.get('emp_name')
        )
        
        def send_whatsapp_notification():
            try:
                emp_phone = manager.get_employee_phone(emp_id)
                if emp_phone:
                    employees = manager.get_employees()
                    emp_info = next((e for e in employees if str(e['emp_id']) == str(emp_id)), {})
                    wa.notify_employee_decision(
                        emp_phone=emp_phone,
                        emp_name=emp_info.get('emp_name', f'Employee #{emp_id}'),
                        manager_name=session.get('emp_name', 'Manager'),
                        request_type=req_type,
                        request_date=req_date,
                        decision=new_status
                    )
            except Exception as wa_err:
                print(f"[WhatsApp] Could not notify employee: {wa_err}")
        
        threading.Thread(target=send_whatsapp_notification, daemon=True).start()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'status': new_status})
        
        flash(f'Request successfully {new_status.lower()}!', 'success')
        return redirect(request.referrer or url_for('manager_dashboard'))
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': str(e)}), 500
        flash(f'Error processing request: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('manager_dashboard'))

@app.route('/cancel-request', methods=['POST'])
def cancel_request():
    if not session.get('logged_in'):
        return redirect(url_for('index'))
        
    emp_id = request.form.get('emp_id')
    req_date = request.form.get('date')
    timestamp = request.form.get('timestamp')
    
    if session.get('user_type') == 'employee' and str(emp_id) != str(session.get('emp_id')):
        flash('You are not authorized to cancel this request.', 'danger')
        return redirect(request.referrer or url_for('employee_dashboard'))
        
    try:
        manager.cancel_request(
            request_date_str=req_date, 
            emp_id=emp_id, 
            timestamp=timestamp, 
            cancelled_by=session.get('emp_name')
        )
        flash('Request successfully cancelled!', 'success')

        try:
            employees = manager.get_employees()
            emp_info = next((e for e in employees if str(e['emp_id']) == str(emp_id)), {})
            emp_team = emp_info.get('emp_team', '')
            req_type_cancelled = request.form.get('type', 'Request')
            if emp_team:
                team_manager = manager.get_manager_for_team(emp_team)
                mgr_phone = team_manager.get('phone', '')
                if mgr_phone and str(emp_id) != str(session.get('emp_id')):
                    wa.notify_manager_request_cancelled(
                        manager_phone=mgr_phone,
                        manager_name=team_manager.get('emp_name', 'Manager'),
                        emp_name=emp_info.get('emp_name', f'Employee #{emp_id}'),
                        request_type=req_type_cancelled,
                        request_date=req_date
                    )
        except Exception as wa_err:
            print(f"[WhatsApp] Could not notify manager of cancellation: {wa_err}")

    except Exception as e:
        flash(f'Error cancelling request: {str(e)}', 'danger')
        
    return redirect(request.referrer or url_for('index'))


# ===== ADMIN ROUTES =====

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        emp_name = request.form.get('emp_name')
        password = request.form.get('password')
        
        try:
            user = manager.authenticate_user(emp_name, password, role='admin')
            if user:
                session['logged_in'] = True
                session['emp_id'] = user['emp_id']
                session['emp_name'] = user['emp_name']
                flash('Login successful', 'success')
                # CEO users (IS_CEO flag) get their own portal
                if user.get('is_ceo'):
                    session['user_type'] = 'ceo'
                    return redirect(url_for('ceo_dashboard'))
                else:
                    session['user_type'] = 'admin'
                    return redirect(url_for('admin_dashboard'))
            else:
                flash('Wrong credentials', 'error')
        except Exception as e:
            flash(f'Login failed: {str(e)}', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin-dashboard')
def admin_dashboard():
    if session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    
    emp_id = session.get('emp_id')
    employees = manager.get_employees()
    _enrich_employees_leave_balances(employees)
    
    # Fail-safe: if user is CEO in database, redirect them to CEO dashboard
    current_emp = next((e for e in employees if str(e.get('emp_id')) == str(emp_id)), {})
    if current_emp.get('is_ceo'):
        session['user_type'] = 'ceo'
        return redirect(url_for('ceo_dashboard'))
        
    today = date.today()
    start_date_str = request.args.get('start_date', today.strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', today.strftime('%Y-%m-%d'))
    
    leave_balance = manager.get_leave_balance_info(emp_id)
    if not isinstance(leave_balance, dict):
        leave_balance = {}

    all_pending_requests = manager.get_pending_requests(req_status='Pending')
    pending_requests = []
    leave_balance_cache = {}
    for req in all_pending_requests:
        if str(req.get('emp_id')) == str(emp_id):
            continue
        req_emp_id = req.get('emp_id')
        req_emp_data = next((e for e in employees if str(e.get('emp_id')) == str(req_emp_id)), {})
        req_emp_team = (req_emp_data.get('emp_team') or '').strip().lower()
        # Admin (Sajeel) only approves Poppi team requests
        if req_emp_team != 'poppi':
            continue
        _attach_request_leave_balance(req, leave_balance_cache)
        req['team'] = req_emp_data.get('emp_team', 'N/A')
        pending_requests.append(req)
    
    all_approved_requests = manager.get_pending_requests(req_status='Approved')
    all_rejected_requests = manager.get_pending_requests(req_status='Rejected')
    recent_requests = []
    
    for req in all_approved_requests + all_rejected_requests:
        if str(req.get('emp_id')) == str(emp_id):
            continue
        req_emp_id = req.get('emp_id')
        req_emp_data = next((e for e in employees if str(e.get('emp_id')) == str(req_emp_id)), {})
        req['team'] = req_emp_data.get('emp_team', 'N/A')
        recent_requests.append(req)
    
    filtered_requests = []
    try:
        sd = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        ed = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        for req in recent_requests:
            req_date_raw = req.get('date')
            if req_date_raw:
                if isinstance(req_date_raw, (date, datetime)):
                    req_date = req_date_raw if isinstance(req_date_raw, date) else req_date_raw.date()
                else:
                    req_date = datetime.strptime(str(req_date_raw).split(' ')[0], '%Y-%m-%d').date()
                if sd <= req_date <= ed:
                    filtered_requests.append(req)
    except Exception as e:
        filtered_requests = recent_requests
    
    filtered_requests.sort(key=lambda x: x.get('date', ''), reverse=True)

    # recent_requests: all teams visible to admin for history
    return render_template('admin_dashboard.html',
                         admin_name=session.get('emp_name'),
                         emp_data={'emp_team': 'Poppi'},
                         leave_balance=leave_balance,
                         requests=pending_requests,
                         recent_requests=filtered_requests,
                         employees=employees,
                         start_date=start_date_str,
                         end_date=end_date_str,
                         today=today.strftime('%Y-%m-%d'))


# ===== CEO DASHBOARD =====

@app.route('/ceo-dashboard')
def ceo_dashboard():
    if session.get('user_type') != 'ceo':
        return redirect(url_for('admin_login'))

    today = date.today()
    start_date_str = request.args.get('start_date', today.strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', today.strftime('%Y-%m-%d'))

    employees = manager.get_employees()
    _enrich_employees_leave_balances(employees)

    # Company-wide approved + rejected requests for history view
    all_approved = manager.get_pending_requests(req_status='Approved')
    all_rejected = manager.get_pending_requests(req_status='Rejected')
    recent_requests = []
    emp_id = session.get('emp_id')
    for req in all_approved + all_rejected:
        req_emp_id = req.get('emp_id')
        req_emp_data = next((e for e in employees if str(e.get('emp_id')) == str(req_emp_id)), {})
        req['team'] = req_emp_data.get('emp_team', 'N/A')
        recent_requests.append(req)

    # Date filter
    filtered_requests = []
    try:
        sd = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        ed = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        for req in recent_requests:
            req_date_raw = req.get('date')
            if req_date_raw:
                if isinstance(req_date_raw, (date, datetime)):
                    req_date = req_date_raw if isinstance(req_date_raw, date) else req_date_raw.date()
                else:
                    req_date = datetime.strptime(str(req_date_raw).split(' ')[0], '%Y-%m-%d').date()
                if sd <= req_date <= ed:
                    filtered_requests.append(req)
    except Exception:
        filtered_requests = recent_requests

    filtered_requests.sort(key=lambda x: x.get('date', ''), reverse=True)

    return render_template('ceo_dashboard.html',
                           ceo_name=session.get('emp_name'),
                           employees=employees,
                           recent_requests=filtered_requests,
                           start_date=start_date_str,
                           end_date=end_date_str,
                           today=today.strftime('%Y-%m-%d'))


@app.route('/admin/performance-report')
def admin_performance_report():
    if session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    
    employees = manager.get_employees()
    _enrich_employees_leave_balances(employees)
    
    employees.sort(key=lambda x: x.get('Remaining_Leaves', 0) or 0, reverse=True)
    return render_template('admin_performance.html', employees=employees)


@app.route('/admin/view-employee/<int:emp_id>')
def admin_view_employee(emp_id):
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))
    
    filter_type = request.args.get('filter_type', 'custom')
    today = date.today()
    
    if filter_type == 'monthly':
        month = int(request.args.get('month', today.month))
        year = int(request.args.get('year', today.year))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year, 12, 31)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    elif filter_type == 'yearly':
        year = int(request.args.get('year', today.year))
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
    
    else:
        start_date_str = request.args.get('start_date', (today - timedelta(days=30)).strftime('%Y-%m-%d'))
        end_date_str = request.args.get('end_date', today.strftime('%Y-%m-%d'))
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
    
    records = manager.get_employee_records(emp_id, start_date, end_date)
    
    employees = manager.get_employees()
    emp_info = next((e for e in employees if str(e['emp_id']) == str(emp_id)), None)

    leave_balance = manager.get_leave_balance_info(emp_id)
    
    return render_template('admin_view_employee.html',
                         employee=emp_info,
                         leave_balance=leave_balance,
                         records=records,
                         filter_type=filter_type,
                         start_date=start_date.strftime('%Y-%m-%d'),
                         end_date=end_date.strftime('%Y-%m-%d'))

@app.route('/logout')
def logout():
    session.clear()
    session.pop('logged_in', None)
    session.pop('user_type', None)
    session.pop('emp_id', None)
    session.pop('emp_name', None)
    session.pop('emp_team', None)
    session.pop('is_manager', None)
    session.pop('is_admin', None)
    response = redirect(url_for('index'))
    response.delete_cookie('session')
    return response

@app.route('/admin/add-employee', methods=['GET', 'POST'])
def add_employee_route():
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))
        
    if request.method == 'POST':
        try:
            emp_name = request.form.get('emp_name')
            emp_team = request.form.get('emp_team')
            role = request.form.get('role')
            contract_type = request.form.get('contract_type')
            contract_start_date = request.form.get('contract_start_date')
            contract_end_date = request.form.get('contract_end_date')
            
            if contract_type == 'Internship':
                total_leaves = 0
            else:
                total_leaves = int(request.form.get('total_leaves', 14))
            
            if not emp_name or not emp_team:
                flash('Name and Team are required', 'error')
                return redirect(url_for('add_employee_route'))
                
            is_admin = (role == 'admin')
            is_manager = (role == 'manager')
            
            manager.add_employee(
                emp_name=emp_name,
                emp_team=emp_team,
                is_admin=is_admin,
                is_manager=is_manager,
                contract_type=contract_type,
                contract_start_date=contract_start_date,
                contract_end_date=contract_end_date,
                total_leaves=total_leaves
            )
            flash(f"Employee '{emp_name}' added successfully!", 'success')
            if session.get('user_type') == 'ceo':
                return redirect(url_for('ceo_dashboard'))
            return redirect(url_for('admin_dashboard'))
            
        except Exception as e:
            flash(f"Error adding employee: {str(e)}", 'error')
            
    return render_template('admin_add_employee.html', admin_name=session.get('emp_name'))

@app.route('/admin/manage-employees')
def manage_employees_route():
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))
    
    try:
        employees = manager.get_employees()
        return render_template('admin_manage_employees.html', employees=employees)
    except Exception as e:
        flash(f"Error loading employees: {str(e)}", 'error')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/edit-employee/<int:emp_id>', methods=['GET', 'POST'])
def edit_employee_route(emp_id):
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))
        
    try:
        employees = manager.get_employees()
        employee = next((e for e in employees if str(e['emp_id']) == str(emp_id)), None)
        
        if not employee:
            flash(f"Employee not found", 'error')
            return redirect(url_for('manage_employees_route'))
            
        if request.method == 'POST':
            emp_name = request.form.get('emp_name') or employee.get('emp_name')
            emp_team = request.form.get('emp_team') or employee.get('emp_team')
            role = request.form.get('role')
            
            if not role:
                if employee.get('is_admin'): role = 'admin'
                elif employee.get('is_manager'): role = 'manager'
                else: role = 'employee'
                
            contract_type = request.form.get('contract_type') or employee.get('Contract_Type')
            contract_start_date = request.form.get('contract_start_date') or employee.get('Contract_Start_Date')
            contract_end_date = request.form.get('contract_end_date') or employee.get('Contract_End_Date')
            
            raw_leaves = request.form.get('total_leaves')
            raw_carried = request.form.get('carried_forward')
            phone = request.form.get('phone', '').strip()
            
            if contract_type == 'Internship':
                total_leaves = 0
                carried_forward = 0
            else:
                total_leaves = int(raw_leaves) if raw_leaves else employee.get('Total_leaves', 14)
                if total_leaves is None: total_leaves = 14
                
                carried_forward = float(raw_carried) if raw_carried else employee.get('Leaves_Carried_Forward', 0)
                if carried_forward is None: carried_forward = 0
            
            is_admin = (role == 'admin')
            is_manager = (role == 'manager')
            
            manager.update_employee(
                emp_id=emp_id,
                emp_name=emp_name,
                emp_team=emp_team,
                is_admin=is_admin,
                is_manager=is_manager,
                contract_type=contract_type,
                contract_start_date=contract_start_date,
                contract_end_date=contract_end_date,
                total_leaves=total_leaves,
                carried_forward=carried_forward,
                phone=phone
            )
            
            flash(f"Employee {emp_name} updated successfully!", 'success')
            return redirect(url_for('manage_employees_route'))
            
        return render_template('admin_edit_employee.html', employee=employee, admin_name=session.get('emp_name'))
        
    except Exception as e:
        flash(f"Error: {str(e)}", 'error')
        return redirect(url_for('manage_employees_route'))

@app.route('/admin/delete-employee/<int:emp_id>', methods=['POST'])
def delete_employee_route(emp_id):
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))
        
    try:
        # Get employee name before deletion for a helpful flash message
        employees = manager.get_employees()
        emp_info = next((e for e in employees if str(e.get('emp_id')) == str(emp_id)), None)
        emp_name = emp_info.get('emp_name', 'Employee') if emp_info else 'Employee'

        manager.delete_employee(emp_id)
        flash(f"Employee '{emp_name}' removed successfully!", 'success')
        
        # Redirection logic back to the previous page or dashboard
        referrer = request.referrer
        if referrer and 'view-employee' not in referrer:
            return redirect(referrer)
            
        if session.get('user_type') == 'ceo':
            return redirect(url_for('ceo_dashboard'))
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        flash(f"Error deleting employee: {str(e)}", 'error')
        return redirect(url_for('admin_view_employee', emp_id=emp_id))


# ===== WHATSAPP ADMIN ROUTES =====

@app.route('/admin/send-daily-summary', methods=['POST'])
def admin_send_daily_summary():
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))

    try:
        _send_daily_summaries()
        flash('✅ Daily attendance summary has been sent to all managers!', 'success')
    except Exception as e:
        flash(f'❌ Error sending daily summary: {str(e)}', 'error')

    return redirect(url_for('admin_dashboard'))


@app.route('/admin/send-absent-alert', methods=['POST'])
def admin_send_absent_alert():
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))

    try:
        _send_absent_alerts()
        flash('✅ Absent employee alert has been sent to all managers!', 'success')
    except Exception as e:
        flash(f'❌ Error sending absent alert: {str(e)}', 'error')

    return redirect(url_for('admin_dashboard'))


@app.route('/admin/whatsapp-settings')
def admin_whatsapp_settings():
    if session.get('user_type') not in ('admin', 'ceo'):
        return redirect(url_for('admin_login'))

    token_set = bool(wa.WHATSAPP_ACCESS_TOKEN)
    phone_id_set = bool(wa.WHATSAPP_PHONE_NUMBER_ID)
    enabled = wa.WHATSAPP_ENABLED
    api_version = wa.GRAPH_API_VERSION

    all_managers = manager.get_all_managers()
    managers_with_phone = [m for m in all_managers if m.get('phone') and str(m.get('phone')).strip() not in ['', 'None', 'nan', 'NaN']]
    managers_without_phone = [m for m in all_managers if m not in managers_with_phone]

    jobs = []
    for job in scheduler.get_jobs():
        jobs.append({
            'id': job.id,
            'name': job.name,
            'next_run': str(job.next_run_time) if job.next_run_time else 'Paused',
            'trigger': str(job.trigger)
        })

    return render_template('admin_whatsapp_settings.html',
                           admin_name=session.get('emp_name'),
                           token_set=token_set,
                           phone_id_set=phone_id_set,
                           enabled=enabled,
                           api_version=api_version,
                           managers_with_phone=managers_with_phone,
                           managers_without_phone=managers_without_phone,
                           scheduled_jobs=jobs)


@app.route('/admin/reset-system-balances')
def reset_system_balances():
    """Route updated to trigger database-only initialization logic"""
    try:
        # Excel logic completely removed. Database balance reset function triggered.
        results = manager.reset_all_balances_in_db()
        return "✅ FINAL RESET SUCCESSFUL IN SNOWFLAKE DATABASE:<br>" + "<br>".join(results)
    except Exception as e:
        return f"❌ Reset Failed: {str(e)}"

@app.route('/save-schedule', methods=['POST'])
def save_schedule():
    if session.get('user_type') not in ['manager', 'admin']:
        return redirect(url_for('employee_login'))
    
    emp_id = session.get('emp_id')
    employees = manager.get_employees()
    current_emp = next((e for e in employees if str(e.get('emp_id')) == str(emp_id)), {})
    if current_emp.get('emp_team', '').lower() != 'overstock' and session.get('user_type') != 'admin':
        flash('Access denied. Only Overstock team can save schedules.', 'error')
        return redirect(url_for('manager_dashboard'))
    
    try:
        valid_from = request.form.get('valid_from')
        valid_until = request.form.get('valid_until')
        
        schedule_data = []
        shift_fields = {
            'shift_weekend_night_emp_id': 'Weekend Night',
            'shift_weekend_morning_emp_id': 'Weekend Morning',
            'shift_night_emp_id': 'Night',
            'shift_morning_emp_id': 'Morning',
            'shift_p1_emp_id': 'Primary (P1)',
            'shift_p2_emp_id': 'Primary (P2)',
            'shift_dev_office_emp_id': 'Development Office',
            'shift_dev_office2_emp_id': 'Development office'
        }
        
        for field_name, shift_name in shift_fields.items():
            emp_id_field = request.form.get(field_name)
            if emp_id_field:
                schedule_data.append({
                    'shift': shift_name,
                    'emp_id': emp_id_field
                })
        
        meeting_lead_week1 = request.form.get('meeting_lead_week1_emp_id')
        meeting_lead_week2 = request.form.get('meeting_lead_week2_emp_id')
        weekly_report_week1 = request.form.get('weekly_report_week1_emp_id')
        weekly_report_week2 = request.form.get('weekly_report_week2_emp_id')
        
        success = manager.save_shift_schedule(
            valid_from, valid_until, schedule_data,
            meeting_lead_week1, meeting_lead_week2,
            weekly_report_week1, weekly_report_week2
        )
        
        if success:
            flash('Schedule saved successfully! (Valid from ' + valid_from + ' to ' + valid_until + ')', 'success')
        else:
            flash('Error saving schedule to database.', 'error')
    except Exception as e:
        flash(f'Error saving schedule: {str(e)}', 'error')
    
    return redirect(url_for('manager_dashboard'))

@app.route('/view-beyond-schedule')
def view_beyond_schedule():
    user_type = session.get('user_type')
    emp_team = (session.get('emp_team') or '').strip().lower()
    
    is_allowed = False
    if user_type in ['admin', 'manager', 'ceo']:
        is_allowed = True
    elif user_type == 'employee' and emp_team == 'overstock':
        is_allowed = True
        
    if not is_allowed:
        return redirect(url_for('employee_login'))
    
    schedules = manager.get_shift_schedules()
    if schedules:
        latest_submitted_schedule = max(schedules, key=lambda s: s.get('submitted_at') or s.get('valid_from'))
        latest_date = latest_submitted_schedule.get('valid_from')
        if latest_date:
            schedules = [s for s in schedules if s.get('valid_from') == latest_date]
            
    employees = manager.get_employees()
    
    # Filter: only display Overstock team members (employees and managers)
    overstock_employees = [e for e in employees if (e.get('emp_team') or '').strip().lower() == 'overstock']
    # Sort managers to the top, then by name
    overstock_employees.sort(key=lambda e: (not e.get('is_manager', False), e.get('emp_name', '').lower()))
    
    # Invert schedules to find who is assigned to what main shift
    emp_to_shift = {}
    meeting_lead_week1 = 'Not Assigned'
    meeting_lead_week2 = 'Not Assigned'
    weekly_report_week1 = 'Not Assigned'
    weekly_report_week2 = 'Not Assigned'
    
    for s in schedules:
        emp_id_str = str(s.get('emp_id'))
        schedule_type = s.get('schedule_type')
        shift_name = s.get('shift_name')
        
        # Resolve names for weekly roles
        resolved_name = next((e.get('emp_name') for e in employees if str(e.get('emp_id')) == emp_id_str), 'Not Assigned')
        
        if schedule_type == 'meeting_lead_week1':
            meeting_lead_week1 = resolved_name
        elif schedule_type == 'meeting_lead_week2':
            meeting_lead_week2 = resolved_name
        elif schedule_type == 'weekly_report_week1':
            weekly_report_week1 = resolved_name
        elif schedule_type == 'weekly_report_week2':
            weekly_report_week2 = resolved_name
        elif schedule_type == 'main':
            emp_to_shift[emp_id_str] = shift_name
            
    # Shift schedules definitions
    SHIFT_SCHEDULES = {
        'Weekend Night': {
            'MON': 'Off Day', 'TUE': 'Off Day', 'WED': 'Weekend Night', 'THU': 'Weekend Night', 
            'FRI': 'Off Day', 'SAT': 'Weekend Night', 'SUN': 'Weekend Night'
        },
        'Weekend Morning': {
            'MON': 'Weekend Morning', 'TUE': 'Weekend Morning', 'WED': 'Off Day', 'THU': 'Off Day', 
            'FRI': 'Off Day', 'SAT': 'Weekend Morning', 'SUN': 'Weekend Morning'
        },
        'Night': {
            'MON': 'Night', 'TUE': 'Night', 'WED': 'Night', 'THU': 'Night', 
            'FRI': 'Night', 'SAT': 'Off Day', 'SUN': 'Off Day'
        },
        'Morning': {
            'MON': 'Morning', 'TUE': 'Morning', 'WED': 'Morning', 'THU': 'Morning', 
            'FRI': 'Morning', 'SAT': 'Off Day', 'SUN': 'Off Day'
        },
        'Primary (P1)': {
            'MON': 'Primary (P1)', 'TUE': 'Primary (P1)', 'WED': 'Primary (P1)', 'THU': 'Primary (P1)', 
            'FRI': 'Primary (P1)', 'SAT': 'Off Day', 'SUN': 'Off Day'
        },
        'Primary (P2)': {
            'MON': 'Primary (P2)', 'TUE': 'Primary (P2)', 'WED': 'Primary (P2)', 'THU': 'Primary (P2)', 
            'FRI': 'Primary (P2)', 'SAT': 'Off Day', 'SUN': 'Off Day'
        },
        'Development Office': {
            'MON': 'Dev 1', 'TUE': 'Dev 1', 'WED': 'Dev 1', 'THU': 'Dev 1', 
            'FRI': 'Dev 1', 'SAT': 'Off Day', 'SUN': 'Off Day'
        },
        'Development office': {
            'MON': 'Dev 2', 'TUE': 'Dev 2', 'WED': 'Dev 2', 'THU': 'Dev 2', 
            'FRI': 'Dev 2', 'SAT': 'Off Day', 'SUN': 'Off Day'
        }
    }
    
    SHIFT_DISPLAY = {
        'Weekend Night': 'Weekend Night',
        'Weekend Morning': 'Weekend Morning',
        'Night': 'Night',
        'Morning': 'Morning',
        'Primary (P1)': 'Primary (P1)',
        'Primary (P2)': 'Primary (P2)',
        'Development Office': 'Dev 1',
        'Development office': 'Dev 2'
    }
    
    grid_data = []
    for emp in overstock_employees:
        emp_id_str = str(emp.get('emp_id'))
        emp_name = emp.get('emp_name')
        is_manager = emp.get('is_manager', False)
        
        assigned_shift = emp_to_shift.get(emp_id_str)
        
        days_schedule = {}
        for d in ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']:
            if assigned_shift and assigned_shift in SHIFT_SCHEDULES:
                days_schedule[d] = SHIFT_SCHEDULES[assigned_shift][d]
            else:
                days_schedule[d] = 'Off Day'
                
        grid_data.append({
            'emp_id': emp.get('emp_id'),
            'emp_name': emp_name,
            'is_manager': is_manager,
            'assigned_shift': assigned_shift or 'Not Assigned',
            'assigned_shift_display': SHIFT_DISPLAY.get(assigned_shift, 'Not Assigned') if assigned_shift else 'Not Assigned',
            'days': days_schedule
        })
        
    return render_template('view_beyond_schedule.html', 
                          admin_name=session.get('emp_name', 'Admin'),
                          grid_data=grid_data,
                          meeting_lead_week1=meeting_lead_week1,
                          meeting_lead_week2=meeting_lead_week2,
                          weekly_report_week1=weekly_report_week1,
                          weekly_report_week2=weekly_report_week2)

@app.route('/export-beyond-schedule')
def export_beyond_schedule():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    user_type = session.get('user_type')
    emp_team = (session.get('emp_team') or '').strip().lower()
    
    is_allowed = False
    if user_type in ['admin', 'manager', 'ceo']:
        is_allowed = True
    elif user_type == 'employee' and emp_team == 'overstock':
        is_allowed = True
        
    if not is_allowed:
        return redirect(url_for('employee_login'))
    
    schedules = manager.get_shift_schedules()
    if schedules:
        latest_submitted_schedule = max(schedules, key=lambda s: s.get('submitted_at') or s.get('valid_from'))
        latest_date = latest_submitted_schedule.get('valid_from')
        if latest_date:
            schedules = [s for s in schedules if s.get('valid_from') == latest_date]
            
    employees = manager.get_employees()
    # Invert schedules to find who is assigned to what main shift
    emp_to_shift = {}
    meeting_lead_week1 = 'Not Assigned'
    meeting_lead_week2 = 'Not Assigned'
    weekly_report_week1 = 'Not Assigned'
    weekly_report_week2 = 'Not Assigned'
    
    for s in schedules:
        emp_id_str = str(s.get('emp_id'))
        schedule_type = s.get('schedule_type')
        shift_name = s.get('shift_name')
        
        resolved_name = next((e.get('emp_name') for e in employees if str(e.get('emp_id')) == emp_id_str), 'Not Assigned')
        
        if schedule_type == 'meeting_lead_week1':
            meeting_lead_week1 = resolved_name
        elif schedule_type == 'meeting_lead_week2':
            meeting_lead_week2 = resolved_name
        elif schedule_type == 'weekly_report_week1':
            weekly_report_week1 = resolved_name
        elif schedule_type == 'weekly_report_week2':
            weekly_report_week2 = resolved_name
        elif schedule_type == 'main':
            emp_to_shift[emp_id_str] = shift_name
            
    # Shift schedules definitions
    SHIFT_SCHEDULES = {
        'Weekend Night': {'MON': 'Off Day', 'TUE': 'Off Day', 'WED': 'Weekend Night', 'THU': 'Weekend Night', 'FRI': 'Off Day', 'SAT': 'Weekend Night', 'SUN': 'Weekend Night'},
        'Weekend Morning': {'MON': 'Weekend Morning', 'TUE': 'Weekend Morning', 'WED': 'Off Day', 'THU': 'Off Day', 'FRI': 'Off Day', 'SAT': 'Weekend Morning', 'SUN': 'Weekend Morning'},
        'Night': {'MON': 'Night', 'TUE': 'Night', 'WED': 'Night', 'THU': 'Night', 'FRI': 'Night', 'SAT': 'Off Day', 'SUN': 'Off Day'},
        'Morning': {'MON': 'Morning', 'TUE': 'Morning', 'WED': 'Morning', 'THU': 'Morning', 'FRI': 'Morning', 'SAT': 'Off Day', 'SUN': 'Off Day'},
        'Primary (P1)': {'MON': 'Primary (P1)', 'TUE': 'Primary (P1)', 'WED': 'Primary (P1)', 'THU': 'Primary (P1)', 'FRI': 'Primary (P1)', 'SAT': 'Off Day', 'SUN': 'Off Day'},
        'Primary (P2)': {'MON': 'Primary (P2)', 'TUE': 'Primary (P2)', 'WED': 'Primary (P2)', 'THU': 'Primary (P2)', 'FRI': 'Primary (P2)', 'SAT': 'Off Day', 'SUN': 'Off Day'},
        'Development Office': {'MON': 'Dev 1', 'TUE': 'Dev 1', 'WED': 'Dev 1', 'THU': 'Dev 1', 'FRI': 'Dev 1', 'SAT': 'Off Day', 'SUN': 'Off Day'},
        'Development office': {'MON': 'Dev 2', 'TUE': 'Dev 2', 'WED': 'Dev 2', 'THU': 'Dev 2', 'FRI': 'Dev 2', 'SAT': 'Off Day', 'SUN': 'Off Day'}
    }
    
    SHIFT_DISPLAY = {
        'Weekend Night': 'Weekend Night', 'Weekend Morning': 'Weekend Morning', 'Night': 'Night', 'Morning': 'Morning',
        'Primary (P1)': 'Primary (P1)', 'Primary (P2)': 'Primary (P2)', 'Development Office': 'Dev 1', 'Development office': 'Dev 2'
    }
    
    # Filter: only display Overstock team members
    overstock_employees = [e for e in employees if (e.get('emp_team') or '').strip().lower() == 'overstock']
    overstock_employees.sort(key=lambda e: (not e.get('is_manager', False), e.get('emp_name', '').lower()))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beyond Schedule"
    ws.views.sheetView[0].showGridLines = True

    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Styles
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_gray_shift = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Headers
    headers = ["EMPLOYEE NAME", "ASSIGNED SHIFT", "MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    ws.append(headers)
    for col_idx in range(1, 10):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx > 2 else align_left
        cell.border = thin_border
        
    font_body = Font(name="Segoe UI", size=10, bold=False, color="000000")
    font_body_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
    font_on = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_dash = Font(name="Segoe UI", size=10, color="94A3B8")
    
    row_idx = 2
    for emp in overstock_employees:
        emp_id_str = str(emp.get('emp_id'))
        emp_name = emp.get('emp_name')
        
        assigned_shift = emp_to_shift.get(emp_id_str)
        assigned_shift_disp = SHIFT_DISPLAY.get(assigned_shift, 'Not Assigned') if assigned_shift else 'Not Assigned'
        
        clean_name = str(emp_name).replace('_', ' ').title()
        ws.cell(row=row_idx, column=1, value=clean_name).font = font_body_bold
        ws.cell(row=row_idx, column=1).alignment = align_left
        ws.cell(row=row_idx, column=1).border = thin_border
        
        ws.cell(row=row_idx, column=2, value=assigned_shift_disp).font = font_body
        ws.cell(row=row_idx, column=2).alignment = align_left
        ws.cell(row=row_idx, column=2).border = thin_border
        
        for d_idx, day_name in enumerate(["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]):
            day_val = 'Off Day'
            if assigned_shift and assigned_shift in SHIFT_SCHEDULES:
                day_val = SHIFT_SCHEDULES[assigned_shift].get(day_name, 'Off Day')
                
            cell_col = 3 + d_idx
            display_val = "Off Day" if day_val == "Off Day" else "ON"
            cell = ws.cell(row=row_idx, column=cell_col, value=display_val)
            cell.alignment = align_center
            cell.border = thin_border
            
            if day_val != "Off Day":
                on_fill = "4F46E5" # Default Night indigo
                if day_val == 'Weekend Night': on_fill = "8B5CF6"
                elif day_val == 'Weekend Morning': on_fill = "F97316"
                elif day_val == 'Night': on_fill = "4F46E5"
                elif day_val == 'Morning': on_fill = "0EA5E9"
                elif day_val == 'Primary (P1)': on_fill = "10B981"
                elif day_val == 'Primary (P2)': on_fill = "059669"
                elif day_val in ['Dev 1', 'Dev 2', 'Dev Office', 'Dev Office 2', 'Development Office', 'Development office']: on_fill = "EC4899"
                
                cell.fill = PatternFill(start_color=on_fill, end_color=on_fill, fill_type="solid")
                cell.font = font_on
            else:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                cell.font = Font(name="Segoe UI", size=9, color="94A3B8", italic=True)
        row_idx += 1

    # Meeting Lead
    ws.cell(row=row_idx, column=1, value="Meeting Lead").font = font_body_bold
    ws.cell(row=row_idx, column=1).alignment = align_left
    ws.cell(row=row_idx, column=1).fill = fill_gray_shift
    ws.cell(row=row_idx, column=1).border = thin_border
    
    cell_day1 = ws.cell(row=row_idx, column=2, value="Monday")
    cell_day1.font = Font(name="Segoe UI", size=10, bold=True, color="1E3A8A")
    cell_day1.alignment = align_center
    cell_day1.fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    cell_day1.border = thin_border
    
    ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=5)
    clean_ml1 = str(meeting_lead_week1).replace('_', ' ').title()
    cell_w1 = ws.cell(row=row_idx, column=3, value=f"W1: {clean_ml1}")
    cell_w1.font = Font(name="Segoe UI", size=10, bold=True, color="1E3A8A")
    cell_w1.alignment = align_center
    
    ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=9)
    clean_ml2 = str(meeting_lead_week2).replace('_', ' ').title()
    cell_w2 = ws.cell(row=row_idx, column=6, value=f"W2: {clean_ml2}")
    cell_w2.font = Font(name="Segoe UI", size=10, bold=True, color="1E3A8A")
    cell_w2.alignment = align_center
    
    for col in range(3, 10):
        ws.cell(row=row_idx, column=col).border = thin_border
        ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    row_idx += 1
    
    # Weekly Report
    ws.cell(row=row_idx, column=1, value="Weekly Report").font = font_body_bold
    ws.cell(row=row_idx, column=1).alignment = align_left
    ws.cell(row=row_idx, column=1).fill = fill_gray_shift
    ws.cell(row=row_idx, column=1).border = thin_border
    
    cell_day2 = ws.cell(row=row_idx, column=2, value="Tuesday")
    cell_day2.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    cell_day2.alignment = align_center
    cell_day2.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    cell_day2.border = thin_border
    
    ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=5)
    clean_wr1 = str(weekly_report_week1).replace('_', ' ').title()
    cell_w1_r = ws.cell(row=row_idx, column=3, value=f"W1: {clean_wr1}")
    cell_w1_r.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    cell_w1_r.alignment = align_center
    
    ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=9)
    clean_wr2 = str(weekly_report_week2).replace('_', ' ').title()
    cell_w2_r = ws.cell(row=row_idx, column=6, value=f"W2: {clean_wr2}")
    cell_w2_r.font = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    cell_w2_r.alignment = align_center
    
    for col in range(3, 10):
        ws.cell(row=row_idx, column=col).border = thin_border
        ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ['A', 'B']:
            max_len = 0
            for cell in col:
                if cell.row < 15:
                    val_str = str(cell.value or '')
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 18)
        else:
            ws.column_dimensions[col_letter].width = 12

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    today = date.today()
    file_name = f"beyond_schedule_{today.strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=file_name
    )

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

if __name__ == '__main__':
    app.run(debug=True, port=5000, use_reloader=False)